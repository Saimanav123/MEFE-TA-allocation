import sys, os
from collections import defaultdict, deque
import itertools

import pandas as pd
import networkx as nx
import matplotlib
matplotlib.use('Qt5Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QTableWidget, QTableWidgetItem,
    QTabWidget, QTextEdit, QFrame, QHeaderView, QMessageBox, QComboBox,
    QCheckBox, QSizePolicy, QButtonGroup, QRadioButton, QToolButton
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor


# ── Palette ───────────────────────────────────────────────────
BG      = "#0f1117"; PANEL  = "#1a1d2e"; CARD   = "#21253a"
ACCENT  = "#6c63ff"; ACCENT2= "#00d4aa"; DANGER = "#ff4757"
TEXT    = "#e8e8f0"; SUBTEXT= "#8888aa"; BORDER = "#2e3150"
GREEN   = "#2ed573"

# ── Graph colours ─────────────────────────────────────────────
COL_COURSE     = "#6c63ff"
COL_TA         = "#00d4aa"
COL_EDGE       = "#555577"
COL_MATCHED    = "#ffd700"
COL_UNASSIGNED = "#ff4757"
COL_CYCLE      = "#ff9f43"
COL_FORCED     = "#54a0ff"

# ── Theorem metadata ──────────────────────────────────────────
THEOREM_LABELS = [
    ("Theorem 4  —  Degree−Capacity ≤ 1  (Poly-time)",          "thm4"),
    ("Theorem 5  —  Capacity=1, Distinct TA Utilities",          "thm5"),
    ("Theorem 6  —  TA Degree = 1  (O(m log m))",               "thm6"),
    ("Theorem 7  —  Constant n & Capacity  (Enumeration)",       "thm7"),
    ("Theorem 9  —  FPT(m)  Brute Force",                        "thm9"),
    ("Theorem 12  —  Exchange Matching  (Existential)",          "thm12"),
    ("Theorem 13  —  Hospital–Residents  (Existential)",         "thm13"),
]

THEOREM_PRECONDITIONS = {
    "thm4": ("Degree − capacity ≤ 1 for every course.  The number of positively-valued "
             "TAs may exceed the required count by at most 1.  Runs in O(m²n²)."),
    "thm5": ("Capacity = 1 for every course.  Each TA must assign strictly distinct "
             "utilities to all their positively-valued courses.  Reduces to Stable "
             "Matching (WSSMTI); runs in O(nm²)."),
    "thm6": ("Every TA may positively value at most one course (TA degree ≤ 1).  "
             "Sorts TAs by grade per course; runs in O(m log m)."),
    "thm7": ("Works on any instance by enumerating all feasible matchings and checking "
             "MEFE.  Tractable only when n (# courses) and capacities are small constants."),
    "thm9": ("Brute-force FPT parameterized by m: tries every possible TA→course "
             "assignment.  Only practical for very small m (≤ 10 TAs).  "
             "Complexity O((n+1)^m)."),
    "thm12": ("TAs should have binary utilities {0, a}.  Course utility = TA grade.  "
              "Hall's condition must hold.  Uses Exchange Matching (Algorithm 4 in paper)."),
    "thm13": ("All TAs and courses must positively value each other.  Each TA must have "
              "distinct utilities for courses; each course must give distinct grades to TAs.  "
              "Reduces to Hospital–Residents; always has a solution under these conditions."),
}


# ═══════════════════════════════════════════════════════════════
#  SHARED HELPERS
# ═══════════════════════════════════════════════════════════════

def read_excel(filepath):
    df_c = pd.read_excel(filepath, sheet_name=0)
    courses    = list(df_c.iloc[:, 0].astype(str))
    capacities = {courses[i]: int(df_c.iloc[i, 1]) for i in range(len(courses))}
    df_t = pd.read_excel(filepath, sheet_name=1)
    tas  = list(df_t.iloc[:, 0].astype(str))

    def load_matrix(idx):
        df = pd.read_excel(filepath, sheet_name=idx, index_col=0)
        df.index   = [str(i) for i in df.index]
        df.columns = [str(c) for c in df.columns]
        return {(r, c): float(df.loc[r, c]) for r in df.index for c in df.columns}

    grade          = load_matrix(2)
    ta_utility     = load_matrix(3)
    course_utility = load_matrix(4)
    df_k = pd.read_excel(filepath, sheet_name=5)
    k    = float(df_k.iloc[0, 0])
    return courses, tas, capacities, grade, ta_utility, course_utility, k


def build_graph(courses, tas, ta_utility):
    """Build neighbourhood dicts: nc[course]=set of TAs, nt[ta]=set of courses."""
    nc, nt = defaultdict(set), defaultdict(set)
    for (ta, c), u in ta_utility.items():
        if u != 0:
            nc[c].add(ta); nt[ta].add(c)
    return dict(nc), dict(nt)


def get_components(courses, tas, nc, nt):
    vc, vt = set(), set()
    comps  = []
    def bfs(start):
        cc, ct = set(), set()
        q = deque([('c', start)])
        while q:
            kind, node = q.popleft()
            if kind == 'c':
                if node in vc: continue
                vc.add(node); cc.add(node)
                for t in nc.get(node, []):
                    if t not in vt: q.append(('t', t))
            else:
                if node in vt: continue
                vt.add(node); ct.add(node)
                for c in nt.get(node, []):
                    if c not in vc: q.append(('c', c))
        return cc, ct
    for c in courses:
        if c not in vc:
            cc, ct = bfs(c)
            if cc: comps.append((cc, ct))
    return comps


def validate(matching, courses, tas, capacities, nc,
             course_utility, ta_utility, grade, k):
    report   = {}
    assigned = defaultdict(list)
    for ta, c in matching.items(): assigned[c].append(ta)

    ok = True; msg = "OK"
    for c in courses:
        if len(assigned.get(c, [])) != capacities[c]:
            ok  = False
            msg = f"'{c}' needs {capacities[c]}, got {len(assigned.get(c,[]))}"
            break
    report['Feasibility'] = {'passed': ok, 'message': msg}

    ok = True; msg = "OK"; avg_utils = {}
    for c in courses:
        tl  = assigned.get(c, [])
        cap = capacities[c]
        if not cap: continue
        avg = sum(course_utility.get((c, t), 0) for t in tl) / cap
        avg_utils[c] = round(avg, 2)
        if avg < k:
            ok = False; msg = f"'{c}' AvgUtil={avg:.2f}<k={k}"; break
    report['Satisfaction'] = {'passed': ok, 'message': msg, 'avg_utils': avg_utils}

    ok = True; msg = "OK"
    for ti in tas:
        ci = matching.get(ti)
        for tj in tas:
            if ti == tj: continue
            cj = matching.get(tj)
            if not cj: continue
            gi_j = grade.get((ti, cj), 0)
            gj_j = grade.get((tj, cj), 0)
            ui_j = ta_utility.get((ti, cj), 0)
            ui_i = ta_utility.get((ti, ci), 0) if ci else 0
            if gi_j >= gj_j and ui_j > ui_i:
                ok = False; msg = f"'{ti}' envies '{tj}' (course '{cj}')"; break
        if not ok: break
    report['Envy-Freeness'] = {'passed': ok, 'message': msg}
    return report


def _build_meta(courses, tas, nc, nt, case_name):
    """Build a meta dict for graph visualisation from connected components."""
    comps = get_components(courses, tas, nc, nt)
    meta  = {}
    for idx, (cc, ct) in enumerate(comps):
        meta[idx] = {
            'case': case_name, 'courses': cc, 'tas': ct,
            'cycle_courses': [], 'cycle_tas': [],
            'unassigned_ta': None, 'forced_courses': set(),
        }
    return meta


# ═══════════════════════════════════════════════════════════════
#  THEOREM 4  (original algorithm)
# ═══════════════════════════════════════════════════════════════

def find_cycle(cc, ct, nc, nt):
    visited, parent = {}, {}
    def nbrs(kind, node):
        return [('t', t) for t in nc.get(node, [])] if kind == 'c' \
               else [('c', c) for c in nt.get(node, [])]
    def dfs(kind, node, pk, pn):
        key = (kind, node)
        visited[key] = True
        for nk, nn in nbrs(kind, node):
            nkey = (nk, nn)
            if nkey == (pk, pn): continue
            if nkey in visited:
                path = [nkey, key]
                cur  = parent.get(key)
                while cur and cur != nkey:
                    path.append(cur); cur = parent.get(cur)
                return path
            parent[nkey] = key
            r = dfs(nk, nn, kind, node)
            if r: return r
        return None
    for kind, node in [('c', c) for c in cc] + [('t', t) for t in ct]:
        key = (kind, node)
        if key not in visited:
            parent[key] = None
            r = dfs(kind, node, None, None)
            if r:
                return [n for (k, n) in r if k == 'c'], \
                       [n for (k, n) in r if k == 't']
    return [], []


def extended_matching(comp_courses, nc, matching, mt, mc):
    changed = True
    while changed:
        changed = False
        for course in comp_courses:
            if course in mc: continue
            nbrs    = nc.get(course, set())
            overlap = nbrs & mt
            if overlap:
                for ta in nbrs - mt:
                    matching[ta] = course; mt.add(ta)
                mc.add(course); changed = True


def solve_theorem4(courses, tas, capacities, grade, ta_utility, course_utility, k):
    log  = []
    nc, nt = build_graph(courses, tas, ta_utility)
    meta   = {}

    for c in courses:
        d, cap = len(nc.get(c, set())), capacities[c]
        log.append(f"Course '{c}': degree={d}, capacity={cap}, diff={d-cap}")
        if d - cap > 1:
            log.append("✗ Theorem 4 does NOT apply (diff > 1)")
            return None, None, log, nc, nt, meta

    comps = get_components(courses, tas, nc, nt)
    log.append(f"\nConnected components: {len(comps)}")
    full_matching = {}

    for idx, (cc, ct) in enumerate(comps):
        log.append(f"\nComponent {idx+1}: courses={sorted(cc)}, TAs={sorted(ct)}")
        E = sum(len(nc.get(c, set())) for c in cc)
        V = len(cc) + len(ct)

        has_zero = any(len(nc.get(c, set())) - capacities[c] == 0 for c in cc)
        has_neg  = any(len(nc.get(c, set())) - capacities[c] <  0 for c in cc)

        if has_neg:
            log.append("  ✗ Not enough TAs")
            return None, None, log, nc, nt, meta

        case = ('case1'   if has_zero else
                'case2_1' if E == V-1 else
                'case2_2' if E == V   else 'case2_3')

        log.append(f"  Case: {case}")
        meta[idx] = {'case': case, 'courses': cc, 'tas': ct,
                     'cycle_courses': [], 'cycle_tas': [],
                     'unassigned_ta': None, 'forced_courses': set()}

        if case == 'case2_3':
            log.append("  ✗ Multiple cycles")
            return None, None, log, nc, nt, meta

        comp_match = {}

        if case == 'case1':
            matching, mt, mc = {}, set(), set()
            forced = set()
            for c in cc:
                if len(nc.get(c, set())) == capacities[c]:
                    for ta in nc.get(c, set()):
                        matching[ta] = c; mt.add(ta)
                    mc.add(c); forced.add(c)
                    log.append(f"  Forced: {c} ← {sorted(nc.get(c, set()))}")
            extended_matching(cc, nc, matching, mt, mc)
            comp_match = matching
            meta[idx]['forced_courses'] = forced

        elif case == 'case2_1':
            found = False
            for unassigned in ct:
                matching, mt, mc = {}, {unassigned}, set()
                extended_matching(cc, nc, matching, mt, mc)
                matching.pop(unassigned, None)
                if len(mc) == len(cc):
                    log.append(f"  TA '{unassigned}' unassigned → valid")
                    comp_match = matching
                    meta[idx]['unassigned_ta'] = unassigned
                    found = True; break
            if not found:
                log.append("  ✗ No valid tree matching")
                return None, None, log, nc, nt, meta

        elif case == 'case2_2':
            cyc_c, cyc_t = find_cycle(cc, ct, nc, nt)
            meta[idx]['cycle_courses'] = cyc_c
            meta[idx]['cycle_tas']     = cyc_t
            log.append(f"  Cycle: {cyc_c}")
            x1 = cyc_c[0] if cyc_c else list(cc)[0]
            nbrs_x1 = list(nc.get(x1, set()))
            found = False
            for excl in nbrs_x1[:2]:
                matching, mt, mc = {}, {excl}, set()
                for ta in nbrs_x1:
                    if ta != excl:
                        matching[ta] = x1; mt.add(ta)
                mc.add(x1)
                extended_matching(cc, nc, matching, mt, mc)
                if len(mc) == len(cc):
                    log.append(f"  Excluded '{excl}' from '{x1}' → valid")
                    comp_match = matching; found = True; break
            if not found:
                log.append("  ✗ No valid cycle matching")
                return None, None, log, nc, nt, meta

        full_matching.update(comp_match)

    report = validate(full_matching, courses, tas, capacities,
                      nc, course_utility, ta_utility, grade, k)
    log.append("\n── VALIDATION ──")
    for key, info in report.items():
        log.append(f"  {'✅' if info['passed'] else '❌'} {key}: {info['message']}")
    ok = all(v['passed'] for v in report.values())
    log.append(f"\n{'✅ VALID MEFE MATCHING FOUND' if ok else '❌ NO VALID MEFE MATCHING'}")
    return (full_matching if ok else None), report, log, nc, nt, meta


# ═══════════════════════════════════════════════════════════════
#  SHARED ALGORITHM HELPERS
# ═══════════════════════════════════════════════════════════════

def _gale_shapley_residents(courses, tas, course_prefs, ta_prefs, capacities):
    """
    Residents-optimal (TAs propose) Gale-Shapley.
    course_prefs[c]  = list of TAs sorted best→worst
    ta_prefs[ta]     = list of courses sorted best→worst
    Returns {ta: course} or None if not all courses can be filled.
    """
    # Pre-compute course rank lookup
    c_rank = {}
    for c in courses:
        prefs = course_prefs.get(c, [])
        c_rank[c] = {ta: i for i, ta in enumerate(prefs)}

    ta_next  = {ta: 0 for ta in tas}
    ta_match = {}                          # ta -> course (tentative)
    held     = defaultdict(list)           # course -> [held tas]

    free_q = deque(ta for ta in tas if ta_prefs.get(ta))

    while free_q:
        ta = free_q.popleft()
        if ta in ta_match:
            continue

        prefs = ta_prefs.get(ta, [])

        # Propose to next acceptable course
        while ta_next[ta] < len(prefs):
            c = prefs[ta_next[ta]]
            ta_next[ta] += 1

            if ta not in c_rank.get(c, {}):
                continue          # course does not accept this TA

            cap    = capacities.get(c, 1)
            h      = held[c]
            h.append(ta)
            h.sort(key=lambda t: c_rank[c].get(t, float('inf')))

            if len(h) <= cap:
                ta_match[ta] = c
                break
            else:
                worst = h.pop()
                if worst != ta:
                    ta_match[ta] = c
                    ta_match.pop(worst, None)
                    free_q.appendleft(worst)
                    break
                # else: ta was immediately rejected → try next proposal

    # Verify every course is fully filled
    for c in courses:
        if len(held[c]) != capacities.get(c, 1):
            return None

    return dict(ta_match)


def _find_feasible_matching(courses, tas, capacities, nc):
    """
    Find any feasible matching via max-flow (networkx).
    Returns {ta: course} or None.
    """
    try:
        G   = nx.DiGraph()
        SRC = '__src__'; SNK = '__snk__'
        G.add_node(SRC); G.add_node(SNK)
        for ta in tas:
            G.add_edge(SRC, ta, capacity=1)
        for c in courses:
            for ta in nc.get(c, set()):
                G.add_edge(ta, c, capacity=1)
            G.add_edge(c, SNK, capacity=capacities.get(c, 1))

        flow_val, flow_dict = nx.maximum_flow(G, SRC, SNK)

        if flow_val < sum(capacities.values()):
            return None

        matching = {}
        for ta in tas:
            for c in courses:
                if flow_dict.get(ta, {}).get(c, 0) > 0:
                    matching[ta] = c
        return matching
    except Exception:
        return None


def _exchange_matching_algo(matching, courses, tas, grade, ta_utility):
    """
    Algorithm 4 (Exchange Matching) from the paper.
    Modifies matching in-place until no unmatched TA envies a matched TA
    based on merit (grade).
    """
    matching = dict(matching)
    tm = set(matching.keys())
    tn = set(tas) - tm

    safety = len(tas) ** 2 + 1
    iters  = 0

    while iters < safety:
        iters += 1
        found  = False

        for ti in list(tn):
            for tj in list(tm):
                cj = matching.get(tj)
                if cj is None:
                    continue
                g_i = grade.get((ti, cj), 0)
                g_j = grade.get((tj, cj), 0)
                u_i = ta_utility.get((ti, cj), 0)
                if g_i > g_j and u_i > 0:
                    # Find TA with lowest grade in cj
                    assigned_cj = [t for t, c in matching.items() if c == cj]
                    t_low = min(assigned_cj, key=lambda t: grade.get((t, cj), 0))
                    # Swap
                    del matching[t_low]
                    matching[ti] = cj
                    tm.discard(t_low); tm.add(ti)
                    tn.add(t_low);     tn.discard(ti)
                    found = True
                    break
            if found:
                break

        if not found:
            break

    return matching


# ═══════════════════════════════════════════════════════════════
#  THEOREM 5  —  Capacity=1, Stable Matching reduction
# ═══════════════════════════════════════════════════════════════

def solve_theorem5(courses, tas, capacities, grade, ta_utility, course_utility, k):
    log  = []
    nc, nt = build_graph(courses, tas, ta_utility)

    # Precondition: capacity = 1 for all courses
    bad = [c for c in courses if capacities[c] != 1]
    if bad:
        log.append(f"✗ Theorem 5 requires capacity=1.  Violating courses: {bad}")
        return None, None, log, nc, nt, {}

    # Precondition: each TA has distinct utilities for positively-valued courses
    for ta in tas:
        utils = [ta_utility.get((ta, c), 0)
                 for c in courses if ta_utility.get((ta, c), 0) > 0]
        if len(utils) != len(set(utils)):
            log.append(f"✗ TA '{ta}' has duplicate utilities for positively-valued courses")
            return None, None, log, nc, nt, {}

    log.append("Preconditions satisfied.")
    log.append("Reducing to Stable Matching (courses=men, TAs=women).")
    log.append("  • Eligible pairs: TA utility > 0  AND  course utility ≥ k")
    log.append("  • Course prefs: by grade (desc);  TA prefs: by utility (strict desc)")

    # Build preference lists restricted to satisfaction-eligible pairs
    ta_prefs = {}
    for ta in tas:
        opts = [(c, ta_utility.get((ta, c), 0))
                for c in courses
                if ta_utility.get((ta, c), 0) > 0
                and course_utility.get((c, ta), 0) >= k]
        opts.sort(key=lambda x: -x[1])
        ta_prefs[ta] = [c for c, _ in opts]

    course_prefs = {}
    for c in courses:
        opts = [(ta, grade.get((ta, c), 0))
                for ta in tas
                if ta_utility.get((ta, c), 0) > 0
                and course_utility.get((c, ta), 0) >= k]
        opts.sort(key=lambda x: -x[1])
        course_prefs[c] = [ta for ta, _ in opts]

    log.append("Running Gale-Shapley (TAs propose, residents-optimal)…")
    matching = _gale_shapley_residents(
        courses, tas, course_prefs, ta_prefs, {c: 1 for c in courses})

    if matching is None:
        log.append("✗ No stable matching that covers all courses with satisfaction ≥ k")
        return None, None, log, nc, nt, {}

    log.append(f"Stable matching: {matching}")
    meta   = _build_meta(courses, tas, nc, nt, 'thm5')
    report = validate(matching, courses, tas, capacities,
                      nc, course_utility, ta_utility, grade, k)
    _log_validation(log, report)
    ok = all(v['passed'] for v in report.values())
    return (matching if ok else None), report, log, nc, nt, meta


# ═══════════════════════════════════════════════════════════════
#  THEOREM 6  —  TA degree = 1
# ═══════════════════════════════════════════════════════════════

def solve_theorem6(courses, tas, capacities, grade, ta_utility, course_utility, k):
    log  = []
    nc, nt = build_graph(courses, tas, ta_utility)

    # Precondition: each TA values ≤ 1 course positively
    bad = [(ta, len(nt.get(ta, set())))
           for ta in tas if len(nt.get(ta, set())) > 1]
    if bad:
        log.append("✗ Theorem 6 requires TA degree ≤ 1.  Violating TAs:")
        for ta, d in bad[:5]:
            log.append(f"    TA '{ta}' has degree {d}")
        if len(bad) > 5:
            log.append(f"    … and {len(bad)-5} more")
        return None, None, log, nc, nt, {}

    log.append("All TA degrees ≤ 1.  Sorting by grade per course…")

    matching = {}
    meta     = {}

    for idx, c in enumerate(courses):
        neighbors = sorted(nc.get(c, set()))
        cap = capacities[c]
        log.append(f"Course '{c}': {len(neighbors)} eligible TAs, capacity={cap}")

        if len(neighbors) < cap:
            log.append(f"  ✗ Not enough TAs for '{c}'")
            return None, None, log, nc, nt, {}

        # Sort descending by grade
        neighbors.sort(key=lambda ta: -grade.get((ta, c), 0))
        assigned   = neighbors[:cap]
        unassigned = neighbors[cap:]

        for ta in assigned:
            matching[ta] = c

        log.append(f"  Assigned   : {sorted(assigned)}")
        if unassigned:
            log.append(f"  Unassigned : {sorted(unassigned)}")

        meta[idx] = {
            'case': 'thm6', 'courses': {c}, 'tas': set(neighbors),
            'cycle_courses': [], 'cycle_tas': [],
            'unassigned_ta': unassigned[0] if unassigned else None,
            'forced_courses': set(),
        }

    report = validate(matching, courses, tas, capacities,
                      nc, course_utility, ta_utility, grade, k)
    _log_validation(log, report)
    ok = all(v['passed'] for v in report.values())
    return (matching if ok else None), report, log, nc, nt, meta


# ═══════════════════════════════════════════════════════════════
#  THEOREM 7  —  Constant n & capacity (Enumeration)
# ═══════════════════════════════════════════════════════════════

def solve_theorem7(courses, tas, capacities, grade, ta_utility, course_utility, k):
    log  = []
    nc, nt = build_graph(courses, tas, ta_utility)

    try:
        from math import comb
        total_combos = 1
        for c in courses:
            e = len(nc.get(c, set()))
            cap = capacities[c]
            total_combos *= comb(e, cap) if e >= cap else 0
    except Exception:
        total_combos = -1

    log.append(f"Theorem 7: Enumerating feasible matchings")
    log.append(f"  Courses={len(courses)}, total capacity={sum(capacities.values())}")
    if total_combos == 0:
        log.append("✗ Some course has fewer eligible TAs than its capacity — no feasible matching.")
        return None, None, log, nc, nt, {}
    if total_combos > 0:
        log.append(f"  Upper-bound combinations: {total_combos:,}")
        if total_combos > 500_000:
            log.append("  ⚠ Large search space — may be slow.")

    course_eligible = {c: sorted(nc.get(c, set())) for c in courses}

    def backtrack(ci, used, assignment):
        if ci == len(courses):
            return dict(assignment)
        c   = courses[ci]
        cap = capacities[c]
        elig = [ta for ta in course_eligible[c] if ta not in used]
        if len(elig) < cap:
            return None
        for combo in itertools.combinations(elig, cap):
            for ta in combo:
                assignment[ta] = c
                used.add(ta)
            result = backtrack(ci + 1, used, assignment)
            if result is not None:
                rep = validate(result, courses, tas, capacities,
                               nc, course_utility, ta_utility, grade, k)
                if all(v['passed'] for v in rep.values()):
                    return result
            for ta in combo:
                assignment.pop(ta, None)
                used.discard(ta)
        return None

    matching = backtrack(0, set(), {})

    if matching is None:
        log.append("✗ No valid MEFE matching found after full enumeration.")
        return None, None, log, nc, nt, {}

    log.append(f"✅ Valid matching found: {matching}")
    meta   = _build_meta(courses, tas, nc, nt, 'thm7')
    report = validate(matching, courses, tas, capacities,
                      nc, course_utility, ta_utility, grade, k)
    _log_validation(log, report)
    ok = all(v['passed'] for v in report.values())
    return (matching if ok else None), report, log, nc, nt, meta


# ═══════════════════════════════════════════════════════════════
#  THEOREM 9  —  FPT(m) Brute Force
# ═══════════════════════════════════════════════════════════════

MAX_BF_COMBOS = 2_000_000

def solve_theorem9(courses, tas, capacities, grade, ta_utility, course_utility, k):
    log  = []
    nc, nt = build_graph(courses, tas, ta_utility)
    m = len(tas); n = len(courses)

    log.append(f"Theorem 9: FPT(m) Brute Force  (m={m}, n={n})")

    # Options per TA: None (unassigned) + courses it values positively
    ta_options = []
    for ta in tas:
        opts = [None] + [c for c in courses if ta_utility.get((ta, c), 0) > 0]
        ta_options.append(opts)

    total = 1
    for opts in ta_options:
        total *= len(opts)
        if total > MAX_BF_COMBOS:
            break

    log.append(f"  Assignments to try: {'>' if total > MAX_BF_COMBOS else ''}{min(total, MAX_BF_COMBOS):,}")

    if total > MAX_BF_COMBOS:
        log.append(f"⚠ Exceeds limit ({MAX_BF_COMBOS:,}).  Aborting to avoid hang.")
        log.append("  Tip: use Theorem 4, 5, 6, or 7 for larger instances.")
        return None, None, log, nc, nt, {}

    checked = 0
    for combo in itertools.product(*ta_options):
        checked += 1
        assignment = {ta: c for ta, c in zip(tas, combo) if c is not None}

        # Quick feasibility check
        counts = defaultdict(int)
        for ta, c in assignment.items():
            counts[c] += 1
        if not all(counts.get(c, 0) == capacities[c] for c in courses):
            continue

        rep = validate(assignment, courses, tas, capacities,
                       nc, course_utility, ta_utility, grade, k)
        if all(v['passed'] for v in rep.values()):
            log.append(f"✅ Valid MEFE matching found (checked {checked:,} assignments)")
            meta = _build_meta(courses, tas, nc, nt, 'thm9')
            _log_validation(log, rep)
            log.append("\n✅ VALID MEFE MATCHING FOUND")
            return assignment, rep, log, nc, nt, meta

    log.append(f"❌ No valid MEFE matching found (checked {checked:,} assignments)")
    return None, None, log, nc, nt, {}


# ═══════════════════════════════════════════════════════════════
#  THEOREM 12  —  Exchange Matching (Existential)
# ═══════════════════════════════════════════════════════════════

def solve_theorem12(courses, tas, capacities, grade, ta_utility, course_utility, k):
    log  = []
    nc, nt = build_graph(courses, tas, ta_utility)

    log.append("Theorem 12: Exchange Matching")
    log.append("Preconditions: binary TA utilities, v_x(t)=grade, Hall's condition")

    total_cap = sum(capacities.values())
    if len(tas) < total_cap:
        log.append(f"✗ Not enough TAs: {len(tas)} available, {total_cap} needed.")
        return None, None, log, nc, nt, {}

    # Step 1: initial feasible matching via max-flow
    log.append("\nStep 1 — Finding initial feasible matching via max-flow…")
    matching = _find_feasible_matching(courses, tas, capacities, nc)

    if matching is None:
        log.append("✗ No feasible matching exists (Hall's condition violated).")
        return None, None, log, nc, nt, {}

    log.append(f"  Initial matching: {matching}")

    # Step 2: exchange to achieve MEFE
    log.append("\nStep 2 — Running Exchange Matching (Algorithm 4)…")
    matching = _exchange_matching_algo(matching, courses, tas, grade, ta_utility)
    log.append(f"  Post-exchange matching: {matching}")

    meta   = _build_meta(courses, tas, nc, nt, 'thm12')
    report = validate(matching, courses, tas, capacities,
                      nc, course_utility, ta_utility, grade, k)
    _log_validation(log, report)
    ok = all(v['passed'] for v in report.values())
    return (matching if ok else None), report, log, nc, nt, meta


# ═══════════════════════════════════════════════════════════════
#  THEOREM 13  —  Hospital–Residents reduction (Existential)
# ═══════════════════════════════════════════════════════════════

def solve_theorem13(courses, tas, capacities, grade, ta_utility, course_utility, k):
    log  = []
    nc, nt = build_graph(courses, tas, ta_utility)

    log.append("Theorem 13: Hospital–Residents (HR) Reduction")
    log.append("Residents = TAs, Hospitals = Courses")

    # Warn about missing positive values
    missing = [(ta, c) for ta in tas for c in courses
               if ta_utility.get((ta, c), 0) == 0]
    if missing:
        log.append(f"  ⚠ {len(missing)} (TA, course) pairs with zero TA-utility "
                   f"(precondition: all should be positive).")

    # TA preference lists: by utility desc (strict for precondition)
    ta_prefs = {}
    for ta in tas:
        opts = [(c, ta_utility.get((ta, c), 0))
                for c in courses if ta_utility.get((ta, c), 0) > 0]
        opts.sort(key=lambda x: -x[1])
        ta_prefs[ta] = [c for c, _ in opts]

    # Course preference lists: by grade desc
    course_prefs = {}
    for c in courses:
        opts = [(ta, grade.get((ta, c), 0))
                for ta in tas if ta_utility.get((ta, c), 0) > 0]
        opts.sort(key=lambda x: -x[1])
        course_prefs[c] = [ta for ta, _ in opts]

    log.append("Running Gale-Shapley (TAs/residents propose)…")
    matching = _gale_shapley_residents(
        courses, tas, course_prefs, ta_prefs, capacities)

    if matching is None:
        log.append("✗ No stable matching found (not enough TAs or Hall's condition violated).")
        return None, None, log, nc, nt, {}

    log.append(f"HR stable matching: {matching}")
    meta   = _build_meta(courses, tas, nc, nt, 'thm13')
    report = validate(matching, courses, tas, capacities,
                      nc, course_utility, ta_utility, grade, k)
    _log_validation(log, report)
    ok = all(v['passed'] for v in report.values())
    return (matching if ok else None), report, log, nc, nt, meta


# ── Log helper ────────────────────────────────────────────────
def _log_validation(log, report):
    log.append("\n── VALIDATION ──")
    for key, info in report.items():
        log.append(f"  {'✅' if info['passed'] else '❌'} {key}: {info['message']}")
    ok = all(v['passed'] for v in report.values())
    log.append(f"\n{'✅ VALID MEFE MATCHING FOUND' if ok else '❌ NO VALID MEFE MATCHING'}")


# ── Dispatcher ───────────────────────────────────────────────
THEOREM_SOLVERS = {
    'thm4':  solve_theorem4,
    'thm5':  solve_theorem5,
    'thm6':  solve_theorem6,
    'thm7':  solve_theorem7,
    'thm9':  solve_theorem9,
    'thm12': solve_theorem12,
    'thm13': solve_theorem13,
}


FRIENDLY_BACKEND_LABELS = {
    'thm4':  "Balanced Matching Engine",
    'thm5':  "Preference-First Engine",
    'thm6':  "Quick Single-Choice Engine",
    'thm7':  "Deep Search Engine",
    'thm9':  "Exhaustive Search Engine",
    'thm12': "Exchange-Based Engine",
    'thm13': "Full Preference Engine",
}

SMART_STRATEGIES = [
    {
        'id': 'auto',
        'title': "Auto Select Best Method (Recommended)",
        'icon': "✅",
        'flag': "Recommended",
        'flag_color': GREEN,
        'tooltip': ("Auto Select: The system studies your file and quietly picks "
                    "the best matching path for your data."),
        'description': ("Best for most users. The system checks your data shape "
                        "and uses the most suitable method automatically."),
    },
    {
        'id': 'fast',
        'title': "Fast Matching (for large data)",
        'icon': "⚡",
        'flag': "Speed",
        'flag_color': ACCENT2,
        'tooltip': ("Fast Matching: Gives quick results when TAs have limited "
                    "choices or the dataset is large."),
        'description': ("Optimized for speed. Useful when many records are loaded "
                        "or each TA applies to only a few courses."),
    },
    {
        'id': 'accurate',
        'title': "High Accuracy Matching",
        'icon': "🎯",
        'flag': "Quality",
        'flag_color': ACCENT,
        'tooltip': ("High Accuracy Matching: Focuses on stronger preference "
                    "alignment when the input structure supports it."),
        'description': ("Puts more emphasis on matching quality and preference "
                        "fit when the uploaded data supports that style."),
    },
    {
        'id': 'compare',
        'title': "Try All Methods (Compare Results)",
        'icon': "🔍",
        'flag': "Heavy",
        'flag_color': DANGER,
        'tooltip': ("Try All Methods: Runs every available matching path, then "
                    "compares the outcomes before choosing the strongest result."),
        'description': ("Runs all available methods and compares the outcomes. "
                        "This is slower, but useful on smaller datasets."),
    },
    {
        'id': 'custom',
        'title': "Custom Selection (Advanced)",
        'icon': "🧪",
        'flag': "Advanced",
        'flag_color': "#f1c40f",
        'tooltip': ("Custom Selection: Lets advanced users choose a matching "
                    "engine manually without exposing technical details."),
        'description': ("Lets you override the recommendation and choose the "
                        "internal engine yourself."),
    },
]

CUSTOM_BACKEND_CHOICES = [
    ("Balanced Matching Engine", "thm4"),
    ("Preference-First Engine", "thm5"),
    ("Quick Single-Choice Engine", "thm6"),
    ("Deep Search Engine", "thm7"),
    ("Exhaustive Search Engine", "thm9"),
    ("Exchange-Based Engine", "thm12"),
    ("Full Preference Engine", "thm13"),
]


def analyze_matching_data(courses, tas, capacities, grade, ta_utility, course_utility):
    """Create user-facing guidance from the uploaded Excel data."""
    nc, nt = build_graph(courses, tas, ta_utility)
    num_courses = len(courses)
    num_tas = len(tas)
    total_slots = sum(capacities.values())
    positive_pairs = sum(
        1 for ta in tas for course in courses if ta_utility.get((ta, course), 0) > 0
    )
    ta_degrees = {ta: len(nt.get(ta, set())) for ta in tas}
    course_degrees = {course: len(nc.get(course, set())) for course in courses}
    avg_ta_degree = (sum(ta_degrees.values()) / num_tas) if num_tas else 0.0
    avg_course_degree = (sum(course_degrees.values()) / num_courses) if num_courses else 0.0
    density = positive_pairs / (num_courses * num_tas) if num_courses and num_tas else 0.0

    all_capacity_one = all(capacities[c] == 1 for c in courses) if courses else False
    ta_degree_le_one = all(deg <= 1 for deg in ta_degrees.values()) if tas else False
    degree_capacity_le_one = all(
        course_degrees[c] - capacities[c] <= 1 for c in courses
    ) if courses else False
    all_courses_fillable = all(
        course_degrees[c] >= capacities[c] for c in courses
    ) if courses else False
    limited_choices = avg_ta_degree <= 2.0
    large_data = num_tas >= 30 or num_courses >= 12 or positive_pairs >= 120
    small_data = num_tas <= 14 and num_courses <= 8 and total_slots <= 10
    very_small_data = num_tas <= 10 and num_courses <= 6 and total_slots <= 8

    positive_ta_utils = [u for u in ta_utility.values() if u > 0]
    binary_ta_utilities = len({u for u in positive_ta_utils}) <= 1 if positive_ta_utils else False
    distinct_positive_ta_utilities = True
    for ta in tas:
        vals = [ta_utility.get((ta, c), 0) for c in courses if ta_utility.get((ta, c), 0) > 0]
        if len(vals) != len(set(vals)):
            distinct_positive_ta_utilities = False
            break

    distinct_course_grades = True
    for c in courses:
        vals = [grade.get((ta, c), 0) for ta in tas if ta_utility.get((ta, c), 0) > 0]
        if len(vals) != len(set(vals)):
            distinct_course_grades = False
            break

    all_positive_pairs = all(
        ta_utility.get((ta, c), 0) > 0 and course_utility.get((c, ta), 0) > 0
        for ta in tas for c in courses
    ) if courses and tas else False

    recommendation = {
        'strategy': 'auto',
        'backend': 'thm4',
        'message': "Auto Select is a safe starting point for this dataset.",
        'why_lines': [
            f"There are {num_courses} courses, {num_tas} TAs, and {total_slots} total slots.",
            "The app can inspect the structure and choose a suitable matching path for you.",
        ],
    }

    if ta_degree_le_one and all_courses_fillable:
        recommendation = {
            'strategy': 'fast',
            'backend': 'thm6',
            'message': "Each TA is connected to at most one course, so Fast Matching is recommended.",
            'why_lines': [
                "Every TA has at most one positive course choice.",
                "That means the app can use a very direct and fast matching path.",
            ],
        }
    elif all_capacity_one and distinct_positive_ta_utilities and all_courses_fillable:
        recommendation = {
            'strategy': 'accurate',
            'backend': 'thm5',
            'message': "All courses have capacity 1, so High Accuracy Matching is recommended.",
            'why_lines': [
                "Each course needs exactly one TA.",
                "TA preferences are well-separated, which is a strong fit for preference-focused matching.",
            ],
        }
    elif all_positive_pairs and distinct_positive_ta_utilities and distinct_course_grades and all_courses_fillable:
        recommendation = {
            'strategy': 'accurate',
            'backend': 'thm13',
            'message': "Preferences are rich on both sides, so High Accuracy Matching is recommended.",
            'why_lines': [
                "Every TA-course pair is positively connected.",
                "Both sides show clear ranking patterns, which helps quality-focused matching.",
            ],
        }
    elif very_small_data:
        recommendation = {
            'strategy': 'compare',
            'backend': 'thm7',
            'message': "Your dataset is small, so comparing all methods is practical and useful.",
            'why_lines': [
                "The instance is small enough to try multiple approaches safely.",
                "Comparing outcomes can help confirm the strongest result.",
            ],
        }
    elif large_data and limited_choices:
        recommendation = {
            'strategy': 'fast',
            'backend': 'thm4' if degree_capacity_le_one else 'thm6',
            'message': "Your data is large and each TA has limited choices, so Fast Matching is recommended.",
            'why_lines': [
                f"The dataset contains {positive_pairs} positive TA-course links.",
                "Preference choices are relatively narrow, which favors quicker methods.",
            ],
        }
    elif degree_capacity_le_one and all_courses_fillable:
        recommendation = {
            'strategy': 'auto',
            'backend': 'thm4',
            'message': "The data has a balanced structure, so Auto Select is recommended.",
            'why_lines': [
                "Course demand is close to course capacity across the instance.",
                "That gives the automatic selector a strong balanced option to start with.",
            ],
        }

    insights = [
        f"Courses: {num_courses}  |  TAs: {num_tas}  |  Total slots: {total_slots}",
        f"Average TA choices: {avg_ta_degree:.1f}  |  Average course reach: {avg_course_degree:.1f}",
        f"Active preference links: {positive_pairs} ({density:.0%} of the matrix)",
    ]

    if all_capacity_one:
        insights.append("All courses have capacity 1.")
    if ta_degree_le_one:
        insights.append("Each TA prefers at most one course.")
    if not all_courses_fillable:
        insights.append("Some courses currently have fewer eligible TAs than required capacity.")

    return {
        'nc': nc,
        'nt': nt,
        'num_courses': num_courses,
        'num_tas': num_tas,
        'total_slots': total_slots,
        'positive_pairs': positive_pairs,
        'avg_ta_degree': avg_ta_degree,
        'avg_course_degree': avg_course_degree,
        'density': density,
        'all_capacity_one': all_capacity_one,
        'ta_degree_le_one': ta_degree_le_one,
        'degree_capacity_le_one': degree_capacity_le_one,
        'all_courses_fillable': all_courses_fillable,
        'limited_choices': limited_choices,
        'large_data': large_data,
        'small_data': small_data,
        'very_small_data': very_small_data,
        'binary_ta_utilities': binary_ta_utilities,
        'distinct_positive_ta_utilities': distinct_positive_ta_utilities,
        'distinct_course_grades': distinct_course_grades,
        'all_positive_pairs': all_positive_pairs,
        'recommended_strategy': recommendation['strategy'],
        'recommended_backend': recommendation['backend'],
        'recommendation_text': recommendation['message'],
        'why_lines': recommendation['why_lines'],
        'insights': insights,
    }


def pick_backend_candidates(strategy_id, analysis, custom_backend=None):
    """Hidden routing from user-friendly strategy names to backend solvers."""
    if strategy_id == 'custom':
        return [custom_backend or analysis.get('recommended_backend', 'thm4')]
    if strategy_id == 'fast':
        if analysis['ta_degree_le_one']:
            return ['thm6', 'thm4', 'thm12']
        return ['thm4', 'thm12', 'thm7']
    if strategy_id == 'accurate':
        if analysis['all_capacity_one'] and analysis['distinct_positive_ta_utilities']:
            return ['thm5', 'thm13', 'thm4']
        if analysis['all_positive_pairs'] and analysis['distinct_positive_ta_utilities']:
            return ['thm13', 'thm5', 'thm4']
        return ['thm13', 'thm5', 'thm4', 'thm7']
    if strategy_id == 'compare':
        return ['thm6', 'thm5', 'thm13', 'thm4', 'thm12', 'thm7', 'thm9']

    candidates = []
    if analysis['ta_degree_le_one']:
        candidates.append('thm6')
    if analysis['all_capacity_one'] and analysis['distinct_positive_ta_utilities']:
        candidates.append('thm5')
    if analysis['all_positive_pairs'] and analysis['distinct_positive_ta_utilities'] and analysis['distinct_course_grades']:
        candidates.append('thm13')
    if analysis['degree_capacity_le_one']:
        candidates.append('thm4')
    if analysis['binary_ta_utilities']:
        candidates.append('thm12')
    if analysis['small_data']:
        candidates.append('thm7')
    if analysis['very_small_data']:
        candidates.append('thm9')
    candidates.extend(['thm4', 'thm12', 'thm7', 'thm9'])

    ordered = []
    for key in candidates:
        if key not in ordered:
            ordered.append(key)
    return ordered


def sanitize_log_lines(lines):
    """Hide theorem and algorithm terminology from the user-facing log."""
    replacements = {
        "Theorem 4, 5, 6, or 7": "the faster balanced or preference-oriented engines",
        "Theorem 4": "Balanced Matching Engine",
        "Theorem 5": "Preference-First Engine",
        "Theorem 6": "Quick Single-Choice Engine",
        "Theorem 7": "Deep Search Engine",
        "Theorem 9": "Exhaustive Search Engine",
        "Theorem 12": "Exchange-Based Engine",
        "Theorem 13": "Full Preference Engine",
        "Stable Matching": "preference-based matching",
        "Gale-Shapley": "the preference alignment step",
        "Hospital–Residents": "the full preference model",
        "Hospital-Residents": "the full preference model",
        "Exchange Matching": "the exchange improvement step",
        "Algorithm 4": "the exchange improvement step",
    }
    cleaned = []
    for line in lines:
        updated = line
        for old, new in replacements.items():
            updated = updated.replace(old, new)
        cleaned.append(updated)
    return cleaned


def execute_strategy(data, strategy_id, analysis, custom_backend=None):
    """Run one or more hidden backend solvers for a user-friendly strategy."""
    candidates = pick_backend_candidates(strategy_id, analysis, custom_backend)
    compare_mode = strategy_id == 'compare'
    attempts = []
    best_result = None

    for backend_key in candidates:
        solver = THEOREM_SOLVERS[backend_key]
        matching, report, log, nc, nt, meta = solver(
            data['courses'], data['tas'], data['capacities'],
            data['grade'], data['ta_utility'], data['course_utility'], data['k']
        )
        passed = bool(matching) and bool(report) and all(info['passed'] for info in report.values())
        checks_passed = sum(1 for info in (report or {}).values() if info['passed'])
        size = len(matching or {})
        attempt = {
            'backend_key': backend_key,
            'backend_label': FRIENDLY_BACKEND_LABELS[backend_key],
            'matching': matching,
            'report': report,
            'log': log,
            'nc': nc,
            'nt': nt,
            'meta': meta,
            'passed': passed,
            'checks_passed': checks_passed,
            'matching_size': size,
        }
        attempts.append(attempt)

        if best_result is None or (
            (attempt['passed'], attempt['checks_passed'], attempt['matching_size']) >
            (best_result['passed'], best_result['checks_passed'], best_result['matching_size'])
        ):
            best_result = attempt

        if passed and not compare_mode:
            best_result = attempt
            break

    summary = []
    summary.append(f"Selected strategy: {next(s['title'] for s in SMART_STRATEGIES if s['id'] == strategy_id)}")
    summary.extend(f"- {line}" for line in analysis.get('insights', []))
    summary.append("")
    summary.append("Recommendation reasoning:")
    summary.extend(f"- {line}" for line in analysis.get('why_lines', []))
    summary.append("")

    if compare_mode:
        summary.append("Comparison summary:")
        for attempt in attempts:
            verdict = "PASS" if attempt['passed'] else "Not suitable"
            summary.append(
                f"- {attempt['backend_label']}: {verdict}, "
                f"assigned {attempt['matching_size']} TA(s), "
                f"checks passed {attempt['checks_passed']}"
            )
        summary.append("")

    if best_result:
        summary.append(f"Chosen engine: {best_result['backend_label']}")
        summary.append("")
        summary.append("Run details:")
        summary.extend(sanitize_log_lines(best_result['log']))

    return best_result, "\n".join(summary)


class SmartSelectionPanel(QWidget):
    """User-friendly strategy chooser that hides theorem names."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.analysis = None
        self.option_meta = {}
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(14)

        title_row = QHBoxLayout()
        title_col = QVBoxLayout()
        title_col.setSpacing(2)
        title_col.addWidget(lbl("Choose Matching Strategy", 16, bold=True))
        intro_lbl = QLabel(
            "The app reads your file, recommends the clearest option, and explains why."
        )
        intro_lbl.setWordWrap(True)
        intro_lbl.setStyleSheet(
            f"color:{SUBTEXT};font-size:11px;background:transparent;"
        )
        title_col.addWidget(intro_lbl)
        title_row.addLayout(title_col)
        title_row.addStretch()

        self.link_why = QToolButton()
        self.link_why.setText("Why this recommendation?")
        self.link_why.setCursor(Qt.PointingHandCursor)
        self.link_why.setStyleSheet(
            f"QToolButton{{color:{ACCENT2};background:{PANEL};border:1px solid {BORDER};"
            f"border-radius:10px;padding:8px 12px;font-size:11px;font-weight:700;}}"
            f"QToolButton:hover{{color:{GREEN};border-color:{ACCENT2};}}"
        )
        self.link_why.clicked.connect(self._toggle_why)
        title_row.addWidget(self.link_why)
        root.addLayout(title_row)

        hero = QFrame()
        hero.setStyleSheet(
            "background:qlineargradient(x1:0,y1:0,x2:1,y2:1,"
            "stop:0 #202744, stop:1 #182234);"
            f"border:1px solid {BORDER};border-radius:16px;"
        )
        hero_lay = QVBoxLayout(hero)
        hero_lay.setContentsMargins(18, 16, 18, 16)
        hero_lay.setSpacing(12)

        self.hero_lbl = QLabel("Upload a file to get a smart recommendation.")
        self.hero_lbl.setWordWrap(True)
        self.hero_lbl.setStyleSheet(
            f"color:{TEXT};font-size:18px;font-weight:800;background:transparent;"
        )
        hero_lay.addWidget(self.hero_lbl)

        self.recommendation_lbl = QLabel(
            "The app will inspect your courses, capacities, and TA choices, then suggest the easiest path."
        )
        self.recommendation_lbl.setWordWrap(True)
        self.recommendation_lbl.setStyleSheet(
            f"background:#163527;color:{GREEN};border:1px solid #275a43;"
            f"border-radius:12px;padding:12px 14px;font-size:12px;font-weight:700;"
        )
        hero_lay.addWidget(self.recommendation_lbl)

        stat_row = QHBoxLayout()
        stat_row.setSpacing(10)
        self.stat_courses = self._make_stat_chip("Courses", "-")
        self.stat_tas = self._make_stat_chip("TAs", "-")
        self.stat_shape = self._make_stat_chip("Data Shape", "Waiting for upload")
        stat_row.addWidget(self.stat_courses)
        stat_row.addWidget(self.stat_tas)
        stat_row.addWidget(self.stat_shape, 1)
        hero_lay.addLayout(stat_row)
        root.addWidget(hero)

        self.why_lbl = QLabel("")
        self.why_lbl.setWordWrap(True)
        self.why_lbl.setVisible(False)
        self.why_lbl.setStyleSheet(
            f"color:{TEXT};background:{PANEL};border:1px dashed {ACCENT2};"
            f"border-radius:12px;padding:12px;font-size:11px;"
        )
        root.addWidget(self.why_lbl)

        self.group = QButtonGroup(self)
        self.group.buttonToggled.connect(self._on_selection_changed)

        for strategy in SMART_STRATEGIES:
            card = QFrame()
            card.setStyleSheet(
                f"background:{CARD};border:1px solid {BORDER};border-radius:16px;"
            )
            lay = QHBoxLayout(card)
            lay.setContentsMargins(16, 14, 16, 14)
            lay.setSpacing(14)

            icon_lbl = QLabel(strategy['icon'])
            icon_lbl.setAlignment(Qt.AlignCenter)
            icon_lbl.setFixedSize(40, 40)
            icon_lbl.setStyleSheet(
                f"background:{PANEL};border:1px solid {BORDER};border-radius:20px;"
                f"font-size:18px;color:{strategy['flag_color']};"
            )
            lay.addWidget(icon_lbl)

            text_col = QVBoxLayout()
            text_col.setSpacing(4)

            radio = QRadioButton(strategy['title'])
            radio.setToolTip(strategy['tooltip'])
            radio.setStyleSheet(
                "QRadioButton{font-size:13px;font-weight:700;background:transparent;}"
                "QRadioButton::indicator{width:18px;height:18px;}"
            )
            self.group.addButton(radio)
            text_col.addWidget(radio)

            desc_lbl = QLabel(strategy['description'])
            desc_lbl.setWordWrap(True)
            desc_lbl.setStyleSheet(
                f"color:{SUBTEXT};font-size:11px;background:transparent;"
            )
            text_col.addWidget(desc_lbl)
            lay.addLayout(text_col, 1)

            right_col = QVBoxLayout()
            right_col.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            right_col.setSpacing(8)

            tip_lbl = QLabel(strategy['tooltip'])
            tip_lbl.setWordWrap(True)
            tip_lbl.setAlignment(Qt.AlignRight)
            tip_lbl.setMaximumWidth(260)
            tip_lbl.setStyleSheet(
                f"color:{SUBTEXT};font-size:10px;background:transparent;"
            )
            right_col.addWidget(tip_lbl)

            flag_lbl = QLabel(strategy['flag'])
            flag_lbl.setStyleSheet(
                f"background:{strategy['flag_color']};color:#0f1117;border-radius:10px;"
                f"padding:4px 10px;font-size:10px;font-weight:800;"
            )
            right_col.addWidget(flag_lbl, 0, Qt.AlignRight)
            lay.addLayout(right_col)

            self.option_meta[strategy['id']] = {
                'card': card,
                'radio': radio,
                'desc': desc_lbl,
                'tip': tip_lbl,
                'strategy': strategy,
            }
            root.addWidget(card)

        self.selection_help_lbl = QLabel(
            "Best for most users. The app checks your data and chooses the smoothest path automatically."
        )
        self.selection_help_lbl.setWordWrap(True)
        self.selection_help_lbl.setStyleSheet(
            f"color:{TEXT};background:{PANEL};border:1px solid {BORDER};"
            f"border-radius:12px;padding:14px;font-size:12px;font-weight:600;"
        )
        root.addWidget(self.selection_help_lbl)

        self.custom_row = QFrame()
        self.custom_row.setVisible(False)
        self.custom_row.setStyleSheet(
            f"background:{PANEL};border:1px solid {BORDER};border-radius:12px;"
        )
        cr = QHBoxLayout(self.custom_row)
        cr.setContentsMargins(12, 10, 12, 10)
        cr.addWidget(lbl("Manual engine:", 11, bold=True))
        self.custom_combo = QComboBox()
        for label, key in CUSTOM_BACKEND_CHOICES:
            self.custom_combo.addItem(label, key)
        self.custom_combo.setToolTip(
            "Advanced override for users who want to force a specific internal engine."
        )
        cr.addWidget(self.custom_combo, 1)
        root.addWidget(self.custom_row)

        self.option_meta['auto']['radio'].setChecked(True)
        self._refresh_styles()

    def _make_stat_chip(self, title, value):
        chip = QFrame()
        chip.setStyleSheet(
            f"background:{PANEL};border:1px solid {BORDER};border-radius:12px;"
        )
        lay = QVBoxLayout(chip)
        lay.setContentsMargins(12, 10, 12, 10)
        lay.setSpacing(2)

        title_lbl = QLabel(title)
        title_lbl.setStyleSheet(
            f"color:{SUBTEXT};font-size:10px;font-weight:700;background:transparent;"
        )
        value_lbl = QLabel(value)
        value_lbl.setWordWrap(True)
        value_lbl.setStyleSheet(
            f"color:{TEXT};font-size:14px;font-weight:800;background:transparent;"
        )
        lay.addWidget(title_lbl)
        lay.addWidget(value_lbl)
        chip.value_lbl = value_lbl
        return chip

    def _toggle_why(self):
        visible = not self.why_lbl.isVisible()
        self.why_lbl.setVisible(visible)
        self.link_why.setText("Hide recommendation details" if visible else "Why this recommendation?")

    def selected_strategy(self):
        for strategy_id, meta in self.option_meta.items():
            if meta['radio'].isChecked():
                return strategy_id
        return 'auto'

    def selected_custom_backend(self):
        return self.custom_combo.currentData()

    def update_analysis(self, analysis):
        self.analysis = analysis
        self.hero_lbl.setText("The app studied your file and found a clear next step.")
        self.recommendation_lbl.setText(f"[Recommended] {analysis['recommendation_text']}")
        self.why_lbl.setText("\n".join(f"- {line}" for line in analysis.get('why_lines', [])))
        self.stat_courses.value_lbl.setText(str(analysis.get('num_courses', "-")))
        self.stat_tas.value_lbl.setText(str(analysis.get('num_tas', "-")))
        shape = []
        if analysis.get('large_data'):
            shape.append("Large")
        elif analysis.get('small_data'):
            shape.append("Compact")
        else:
            shape.append("Medium")
        if analysis.get('ta_degree_le_one'):
            shape.append("Simple choices")
        elif analysis.get('limited_choices'):
            shape.append("Limited choices")
        else:
            shape.append("Rich preferences")
        self.stat_shape.value_lbl.setText(" | ".join(shape))
        recommended = analysis.get('recommended_strategy', 'auto')
        if recommended in self.option_meta:
            self.option_meta[recommended]['radio'].setChecked(True)
        self._refresh_styles()
        return
        self.recommendation_lbl.setText(f"💡 {analysis['recommendation_text']}")
        self.why_lbl.setText("\n".join(f"• {line}" for line in analysis.get('why_lines', [])))
        recommended = analysis.get('recommended_strategy', 'auto')
        if recommended in self.option_meta:
            self.option_meta[recommended]['radio'].setChecked(True)
        self._refresh_styles()

    def _on_selection_changed(self, *_args):
        strategy_id = self.selected_strategy()
        if hasattr(self, 'selection_help_lbl'):
            self.selection_help_lbl.setText(
                self.option_meta[strategy_id]['strategy']['description']
            )
        if hasattr(self, 'custom_row'):
            self.custom_row.setVisible(strategy_id == 'custom')
        self._refresh_styles()

    def _refresh_styles(self):
        recommended = self.analysis.get('recommended_strategy') if self.analysis else 'auto'
        selected = self.selected_strategy()
        for strategy_id, meta in self.option_meta.items():
            border = BORDER
            bg = CARD
            desc_color = SUBTEXT
            if strategy_id == recommended:
                border = GREEN
                bg = "#173427"
                desc_color = "#b9e7c9"
            if strategy_id == selected:
                border = ACCENT if strategy_id != recommended else GREEN
                bg = "#252d4c" if strategy_id != recommended else "#173427"
                desc_color = "#dbe1ff" if strategy_id != recommended else "#d4f4de"
            if strategy_id == 'compare':
                meta['card'].setToolTip("This option may take longer because it compares multiple methods.")
            meta['card'].setStyleSheet(
                f"background:{bg};border:2px solid {border};border-radius:16px;"
            )
            meta['desc'].setStyleSheet(
                f"color:{desc_color};font-size:11px;background:transparent;"
            )
            meta['tip'].setStyleSheet(
                f"color:{desc_color};font-size:10px;background:transparent;"
            )
        return
        for strategy_id, meta in self.option_meta.items():
            border = BORDER
            bg = CARD
            if strategy_id == recommended:
                border = GREEN
                bg = "#183324"
            if strategy_id == selected:
                border = ACCENT if strategy_id != recommended else GREEN
                bg = "#252b44" if strategy_id != recommended else "#183324"
            if strategy_id == 'compare':
                meta['card'].setToolTip("This option may take longer because it compares multiple methods.")
            meta['card'].setStyleSheet(
                f"background:{bg};border:2px solid {border};border-radius:12px;"
            )


# ═══════════════════════════════════════════════════════════════
#  GRAPH CANVAS
# ═══════════════════════════════════════════════════════════════

class GraphCanvas(FigureCanvas):
    def __init__(self, parent=None):
        self.fig = Figure(facecolor='#0f1117', tight_layout=False)
        super().__init__(self.fig)
        self.setParent(parent)
        self._placeholder()

    def _placeholder(self):
        self.fig.clear()
        ax = self.fig.add_subplot(111)
        ax.set_facecolor('#0f1117')
        ax.text(0.5, 0.5,
                'Upload Excel  →  Run Matching\nto visualize the Bipartite Graph',
                ha='center', va='center', fontsize=15,
                color='#8888aa', transform=ax.transAxes,
                fontfamily='monospace')
        ax.axis('off')
        self.draw()

    def draw_graph(self, courses, tas, capacities, nc,
                   matching, meta, show_weights,
                   ta_utility, course_utility, grade, weight_mode):
        self.fig.clear()
        comps = list(meta.values())
        if not comps:
            self._placeholder(); return

        n     = len(comps)
        ncols = min(n, 2)
        nrows = (n + 1) // 2
        axes  = self.fig.subplots(nrows, ncols, squeeze=False)

        for idx, m in enumerate(comps):
            ax = axes[idx // ncols][idx % ncols]
            ax.set_facecolor('#13162a')
            self._draw_comp(ax, m, nc, capacities, matching,
                            show_weights, ta_utility, course_utility,
                            grade, weight_mode, idx + 1)

        for idx in range(n, nrows * ncols):
            axes[idx // ncols][idx % ncols].set_visible(False)

        self._legend()
        self.fig.subplots_adjust(hspace=0.4, wspace=0.35, bottom=0.12, top=0.95)
        self.draw()

    def _draw_comp(self, ax, m, nc, capacities, matching,
                   show_weights, ta_u, cu, grade, wmode, comp_idx):
        cc     = m['courses']
        ct     = m['tas']
        case   = m['case']
        cyc_c  = set(m.get('cycle_courses', []))
        cyc_t  = set(m.get('cycle_tas',     []))
        forced = m.get('forced_courses', set())
        unass  = m.get('unassigned_ta')

        cc_list = sorted(cc)
        ct_list = sorted(ct)

        G = nx.Graph()
        for c in cc_list: G.add_node(c, kind='course')
        for t in ct_list: G.add_node(t, kind='ta')
        for c in cc_list:
            for t in nc.get(c, []):
                if t in ct: G.add_edge(c, t)

        pos = {}
        for i, c in enumerate(cc_list):
            y = 0 if len(cc_list) == 1 else 1 - i * 2.0 / (len(cc_list) - 1)
            pos[c] = (-1.0, y)
        for i, t in enumerate(ct_list):
            y = 0 if len(ct_list) == 1 else 1 - i * 2.0 / (len(ct_list) - 1)
            pos[t] = (1.0, y)

        node_col  = []
        node_size = []
        for node in G.nodes():
            kind = G.nodes[node]['kind']
            if kind == 'course':
                col = (COL_FORCED if node in forced else
                       COL_CYCLE  if node in cyc_c  else COL_COURSE)
                node_size.append(2000)
            else:
                col = (COL_UNASSIGNED if node == unass else
                       COL_CYCLE       if node in cyc_t  else COL_TA)
                node_size.append(1600)
            node_col.append(col)

        matched_pairs = set()
        if matching:
            for ta, c in matching.items():
                if ta in ct and c in cc:
                    matched_pairs.add((c, ta)); matched_pairs.add((ta, c))

        solid, dashed = [], []
        for u, v in G.edges():
            if (u, v) in matched_pairs or (v, u) in matched_pairs:
                solid.append((u, v))
            else:
                dashed.append((u, v))

        nx.draw_networkx_nodes(G, pos, ax=ax,
                               node_color=node_col, node_size=node_size,
                               edgecolors='#ffffff', linewidths=1.8)
        if solid:
            nx.draw_networkx_edges(G, pos, edgelist=solid, ax=ax,
                                   edge_color=COL_MATCHED, width=3.5,
                                   style='solid', alpha=0.95)
        if dashed:
            nx.draw_networkx_edges(G, pos, edgelist=dashed, ax=ax,
                                   edge_color=COL_EDGE, width=1.2,
                                   style='dashed', alpha=0.55)

        nx.draw_networkx_labels(G, pos, ax=ax,
                                font_size=9, font_color='white',
                                font_weight='bold')

        for c in cc_list:
            x, y = pos[c]
            ax.annotate(f"cap={capacities.get(c,'?')}",
                        xy=(x, y), xytext=(x - 0.38, y),
                        fontsize=7, color='#aaaacc',
                        ha='right', va='center')

        if show_weights:
            elabels = {}
            for u, v in G.edges():
                c_n = u if G.nodes[u]['kind'] == 'course' else v
                t_n = v if G.nodes[u]['kind'] == 'course' else u
                if   wmode == 'TA Utility':
                    elabels[(u, v)] = f"u={ta_u.get((t_n,c_n),0)}"
                elif wmode == 'Course Utility':
                    elabels[(u, v)] = f"v={cu.get((c_n,t_n),0)}"
                elif wmode == 'Grade':
                    elabels[(u, v)] = f"g={grade.get((t_n,c_n),0)}"
                else:
                    elabels[(u, v)] = (f"u={ta_u.get((t_n,c_n),0)}\n"
                                       f"v={cu.get((c_n,t_n),0)}")
            nx.draw_networkx_edge_labels(
                G, pos, elabels, ax=ax,
                font_size=7, font_color='#ffdd88', label_pos=0.35,
                bbox=dict(boxstyle='round,pad=0.2', facecolor='#1a1d2e',
                          edgecolor='none', alpha=0.85))

        ylim = ax.get_ylim()
        top  = ylim[1] if ylim[1] != ylim[0] else 1.3
        ax.text(-1.0, top + 0.08, 'COURSES', ha='center', va='bottom',
                fontsize=9, color=COL_COURSE, fontweight='bold')
        ax.text( 1.0, top + 0.08, 'TAs', ha='center', va='bottom',
                fontsize=9, color=COL_TA, fontweight='bold')

        case_label = {
            'case1':   '⚡ Case 1 — Forced',
            'case2_1': '🌲 Case 2.1 — Tree',
            'case2_2': '🔄 Case 2.2 — Cycle',
            'case2_3': '❌ Case 2.3 — Multi-cycle',
            'thm5':    '🔗 Preference-First Engine',
            'thm6':    '📊 Quick Single-Choice Engine',
            'thm7':    '🔢 Deep Search Engine',
            'thm9':    '💪 Exhaustive Search Engine',
            'thm12':   '🔄 Exchange-Based Engine',
            'thm13':   '🏥 Full Preference Engine',
        }.get(case, case)

        ax.set_title(f"Component {comp_idx}  |  {case_label}",
                     color='white', fontsize=11, fontweight='bold', pad=14)
        ax.set_xlim(-1.85, 1.85)
        ax.axis('off')

    def _legend(self):
        patches = [
            mpatches.Patch(color=COL_COURSE,     label='Course'),
            mpatches.Patch(color=COL_TA,         label='TA'),
            mpatches.Patch(color=COL_FORCED,     label='Forced (Case 1)'),
            mpatches.Patch(color=COL_CYCLE,      label='Cycle node'),
            mpatches.Patch(color=COL_UNASSIGNED, label='Unassigned TA'),
            mpatches.Patch(color=COL_MATCHED,    label='Matched edge'),
            mpatches.Patch(color=COL_EDGE,       label='Unmatched edge'),
        ]
        self.fig.legend(handles=patches, loc='lower center', ncol=4,
                        frameon=True, facecolor='#21253a', edgecolor=BORDER,
                        labelcolor='white', fontsize=9,
                        bbox_to_anchor=(0.5, 0.0))


# ═══════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════

def lbl(text, size=13, bold=False, color=TEXT):
    w = QLabel(text)
    w.setStyleSheet(f"color:{color};font-size:{size}px;"
                    f"{'font-weight:bold;' if bold else ''}background:transparent;")
    return w


def make_btn(text, bg=ACCENT, hover=None):
    hover = hover or bg
    b = QPushButton(text)
    b.setCursor(Qt.PointingHandCursor)
    b.setStyleSheet(f"""
        QPushButton{{background:{bg};color:#fff;border:none;
            border-radius:8px;padding:10px 22px;
            font-size:13px;font-weight:600;}}
        QPushButton:hover{{background:{hover};}}
        QPushButton:disabled{{background:#333;color:#666;}}""")
    return b


def card_frame():
    f = QFrame()
    f.setStyleSheet(
        f"background:{CARD};border-radius:12px;border:1px solid {BORDER};")
    return f


def make_table(headers):
    t = QTableWidget(0, len(headers))
    t.setHorizontalHeaderLabels(headers)
    t.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    t.verticalHeader().setVisible(False)
    t.setEditTriggers(QTableWidget.NoEditTriggers)
    t.setAlternatingRowColors(True)
    t.setStyleSheet(
        t.styleSheet() +
        "QTableWidget{alternate-background-color:#1e2235;}")
    return t


def add_row(tbl, vals, colors=None):
    r = tbl.rowCount(); tbl.insertRow(r)
    for col, val in enumerate(vals):
        item = QTableWidgetItem(str(val))
        item.setTextAlignment(Qt.AlignCenter)
        if colors and col < len(colors) and colors[col]:
            item.setForeground(QColor(colors[col]))
        tbl.setItem(r, col, item)


# ═══════════════════════════════════════════════════════════════
#  MAIN WINDOW
# ═══════════════════════════════════════════════════════════════

class MEFEApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("MEFE Matching  |  Graph Visualization")
        self.setMinimumSize(1200, 800)
        self.resize(1400, 920)
        self._data = {}
        self._apply_style()
        self._build_ui()

    # ── Stylesheet ───────────────────────────────────────────
    def _apply_style(self):
        self.setStyleSheet(f"""
            QMainWindow,QWidget{{background:{BG};color:{TEXT};
                font-family:'Segoe UI',sans-serif;}}
            QTabWidget::pane{{border:1px solid {BORDER};border-radius:8px;
                background:{PANEL};}}
            QTabBar::tab{{background:{CARD};color:{SUBTEXT};padding:10px 20px;
                border-top-left-radius:8px;border-top-right-radius:8px;
                font-size:13px;}}
            QTabBar::tab:selected{{background:{ACCENT};color:#fff;font-weight:bold;}}
            QTableWidget{{background:{PANEL};color:{TEXT};gridline-color:{BORDER};
                border:none;border-radius:8px;font-size:12px;}}
            QTableWidget::item:selected{{background:{ACCENT};}}
            QHeaderView::section{{background:{CARD};color:{SUBTEXT};border:none;
                padding:6px;font-weight:bold;font-size:12px;}}
            QTextEdit{{background:{PANEL};color:{TEXT};border:1px solid {BORDER};
                border-radius:8px;font-family:'Consolas',monospace;
                font-size:12px;padding:8px;}}
            QComboBox{{background:{CARD};color:{TEXT};border:1px solid {BORDER};
                border-radius:6px;padding:6px 12px;font-size:12px;min-width:120px;}}
            QComboBox::drop-down{{border:none;}}
            QComboBox QAbstractItemView{{background:{CARD};color:{TEXT};
                selection-background-color:{ACCENT};}}
            QCheckBox{{color:{TEXT};font-size:12px;spacing:6px;}}
            QCheckBox::indicator{{width:16px;height:16px;border-radius:4px;
                border:1px solid {BORDER};background:{PANEL};}}
            QCheckBox::indicator:checked{{background:{ACCENT};border-color:{ACCENT};}}
            QScrollBar:vertical{{background:{PANEL};width:8px;border-radius:4px;}}
            QScrollBar::handle:vertical{{background:{ACCENT};border-radius:4px;}}
        """)

    # ── Scaffold ─────────────────────────────────────────────
    def _build_ui(self):
        root = QWidget(); self.setCentralWidget(root)
        main = QVBoxLayout(root)
        main.setContentsMargins(20, 20, 20, 20); main.setSpacing(14)

        # Header
        hdr = QHBoxLayout()
        hdr.addWidget(lbl("MEFE Matching", 22, bold=True, color=ACCENT))
        self.subtitle_lbl = lbl(
            "Guided strategy selection for TA-course allocation",
            13, color=SUBTEXT)
        hdr.addWidget(self.subtitle_lbl)
        hdr.addStretch()
        main.addLayout(hdr)

        self.tabs = QTabWidget()
        self.tabs.addTab(self._tab_upload(),  "📂  Upload")
        self.tabs.addTab(self._tab_results(), "⚙️  Results")
        self.tabs.addTab(self._tab_graph(),   "🕸️  Graph")
        self.tabs.addTab(self._tab_log(),     "📋  Log")
        main.addWidget(self.tabs)

    # ══════════════════════════════════════════════════════
    #  TAB 1 — UPLOAD
    # ══════════════════════════════════════════════════════
    def _tab_upload(self):
        w = QWidget(); lay = QVBoxLayout(w); lay.setSpacing(14)

        zone = card_frame()
        zl   = QVBoxLayout(zone); zl.setAlignment(Qt.AlignCenter)
        zl.setContentsMargins(40, 35, 40, 35); zl.setSpacing(10)

        ic = lbl("📂", 46, color=ACCENT); ic.setAlignment(Qt.AlignCenter)
        zl.addWidget(ic)
        zl.addWidget(lbl("Upload Excel Input File", 16, bold=True))
        zl.addWidget(lbl("6 sheets: Courses · TAs · Grade_g · "
                         "TA_Utility_u · Course_Utility_v · Threshold_k",
                         12, color=SUBTEXT))

        br = QHBoxLayout()
        self.btn_browse = make_btn("  Browse File  ", ACCENT)
        self.btn_sample = make_btn("  Download Sample  ", ACCENT2)
        self.btn_browse.clicked.connect(self._browse)
        self.btn_sample.clicked.connect(self._make_sample)
        br.addStretch()
        br.addWidget(self.btn_browse); br.addWidget(self.btn_sample)
        br.addStretch()
        zl.addLayout(br)

        self.file_lbl = lbl("No file selected", 12, color=SUBTEXT)
        self.file_lbl.setAlignment(Qt.AlignCenter)
        zl.addWidget(self.file_lbl)
        lay.addWidget(zone)

        self._sc = {}
        sc_row   = QHBoxLayout()
        for name in ["Courses", "TAs", "Grade", "TA Utility", "Course Util", "k"]:
            f  = card_frame(); f.setFixedHeight(78)
            vl = QVBoxLayout(f); vl.setAlignment(Qt.AlignCenter)
            t  = lbl(name, 10, color=SUBTEXT); t.setAlignment(Qt.AlignCenter)
            v  = lbl("—", 15, bold=True, color=ACCENT2)
            v.setAlignment(Qt.AlignCenter); v.setObjectName("val")
            vl.addWidget(t); vl.addWidget(v)
            self._sc[name] = f; sc_row.addWidget(f)
        lay.addLayout(sc_row)

        self.prev_tabs = QTabWidget(); self.prev_tabs.setVisible(False)
        lay.addWidget(self.prev_tabs)
        return w

    def _upd_sc(self, name, val, ok=True):
        v = self._sc[name].findChild(QLabel, "val")
        if v:
            v.setText(str(val))
            c = GREEN if ok else DANGER
            v.setStyleSheet(f"color:{c};font-size:15px;"
                            f"font-weight:bold;background:transparent;")

    # ══════════════════════════════════════════════════════
    #  TAB 2 — RESULTS
    # ══════════════════════════════════════════════════════
    def _tab_results(self):
        w = QWidget(); lay = QVBoxLayout(w); lay.setSpacing(12)

        # ── Theorem selector ──────────────────────────────
        sel_card = card_frame()
        sl       = QVBoxLayout(sel_card)
        sl.setContentsMargins(16, 12, 16, 12); sl.setSpacing(8)

        top_row = QHBoxLayout()
        top_row.addWidget(lbl("Matching plan:", 12, bold=True))
        self.strategy_status_lbl = lbl(
            "Upload your Excel file to unlock guided suggestions.",
            11, color=SUBTEXT)
        self.strategy_status_lbl.setWordWrap(True)
        top_row.addWidget(self.strategy_status_lbl)
        top_row.addStretch()

        self.btn_run = make_btn("  Run Matching  ", ACCENT)
        self.btn_run.setFixedHeight(40)
        self.btn_run.setEnabled(False)
        self.btn_run.clicked.connect(self._run)
        top_row.addWidget(self.btn_run)
        sl.addLayout(top_row)

        self.smart_panel = SmartSelectionPanel()
        sl.addWidget(self.smart_panel)

        helper_lbl = QLabel(
            "What happens next: upload your Excel file, review the recommended strategy, then click Run Matching."
        )
        helper_lbl.setWordWrap(True)
        helper_lbl.setStyleSheet(
            f"color:{SUBTEXT};font-size:11px;background:{PANEL};"
            f"border:1px solid {BORDER};border-radius:10px;padding:10px;"
        )
        sl.addWidget(helper_lbl)
        lay.addWidget(sel_card)

        # ── Verdict banner ───────────────────────────────
        self.verdict = QLabel("Upload an Excel file and the app will recommend the best matching strategy.")
        self.verdict.setAlignment(Qt.AlignCenter)
        self.verdict.setStyleSheet(
            f"background:{CARD};color:{SUBTEXT};border-radius:10px;padding:14px;"
            f"font-size:15px;font-weight:600;border:1px solid {BORDER};")
        lay.addWidget(self.verdict)

        # ── Result tables ────────────────────────────────
        panels = QHBoxLayout(); panels.setSpacing(14)

        lc = QVBoxLayout()
        lc.addWidget(lbl("Matching Result", 13, bold=True))
        self.match_tbl = make_table(["TA", "Assigned Course"])
        lc.addWidget(self.match_tbl); panels.addLayout(lc, 1)

        rc = QVBoxLayout()
        rc.addWidget(lbl("Validity Checks", 13, bold=True))
        self.chk_tbl = make_table(["Check", "Result", "Details"])
        rc.addWidget(self.chk_tbl); panels.addLayout(rc, 1)

        lay.addLayout(panels)
        lay.addWidget(lbl("Course Summary", 13, bold=True))
        self.course_tbl = make_table(
            ["Course", "Capacity", "Assigned TAs", "AvgUtil", "≥ k"])
        lay.addWidget(self.course_tbl)

        exp_row = QHBoxLayout()
        self.btn_exp = make_btn("  💾  Export to Excel  ", ACCENT2)
        self.btn_exp.setEnabled(False)
        self.btn_exp.clicked.connect(self._export)
        exp_row.addStretch(); exp_row.addWidget(self.btn_exp); exp_row.addStretch()
        lay.addLayout(exp_row)
        return w

    def _on_theorem_change(self, idx):
        self.subtitle_lbl.setText("Guided strategy selection for TA-course allocation")

    # ══════════════════════════════════════════════════════
    #  TAB 3 — GRAPH
    # ══════════════════════════════════════════════════════
    def _tab_graph(self):
        w = QWidget(); lay = QVBoxLayout(w); lay.setSpacing(10)

        ctrl = QHBoxLayout(); ctrl.setSpacing(16)

        self.chk_weights = QCheckBox("Show edge weights")
        self.chk_weights.setChecked(True)
        self.chk_weights.stateChanged.connect(self._refresh_graph)
        ctrl.addWidget(self.chk_weights)

        ctrl.addWidget(lbl("Weight:", 12, color=SUBTEXT))
        self.cmb_weight = QComboBox()
        self.cmb_weight.addItems(
            ["TA Utility  (u)", "Course Utility  (v)", "Grade  (g)", "Both u & v"])
        self.cmb_weight.currentIndexChanged.connect(self._refresh_graph)
        ctrl.addWidget(self.cmb_weight)
        ctrl.addStretch()

        ctrl.addWidget(lbl("View:", 12, color=SUBTEXT))
        self.cmb_view = QComboBox()
        self.cmb_view.addItems(
            ["After Matching  (show result)", "Before Matching  (graph only)"])
        self.cmb_view.currentIndexChanged.connect(self._refresh_graph)
        ctrl.addWidget(self.cmb_view)

        btn_ref = make_btn("  🔄 Refresh  ", ACCENT)
        btn_ref.setFixedHeight(36)
        btn_ref.clicked.connect(self._refresh_graph)
        ctrl.addWidget(btn_ref)
        lay.addLayout(ctrl)

        leg = QHBoxLayout(); leg.setSpacing(14)
        for color, text in [
            (COL_COURSE,     "Course"), (COL_TA,         "TA"),
            (COL_FORCED,     "Forced (Case 1)"), (COL_CYCLE, "Cycle node"),
            (COL_UNASSIGNED, "Unassigned TA"),
            (COL_MATCHED,    "Matched ══"), (COL_EDGE, "Unmatched - -"),
        ]:
            dot = QLabel("●")
            dot.setStyleSheet(
                f"color:{color};font-size:16px;background:transparent;")
            leg.addWidget(dot)
            leg.addWidget(lbl(text, 11, color=SUBTEXT))
        leg.addStretch()
        lay.addLayout(leg)

        self.canvas = GraphCanvas()
        lay.addWidget(self.canvas, stretch=1)

        self.graph_info = QLabel("Run matching to populate the graph.")
        self.graph_info.setAlignment(Qt.AlignCenter)
        self.graph_info.setStyleSheet(
            f"color:{SUBTEXT};font-size:11px;background:{CARD};"
            f"border-radius:8px;padding:8px;border:1px solid {BORDER};")
        lay.addWidget(self.graph_info)
        return w

    # ══════════════════════════════════════════════════════
    #  TAB 4 — LOG
    # ══════════════════════════════════════════════════════
    def _tab_log(self):
        w = QWidget(); lay = QVBoxLayout(w)
        lay.addWidget(lbl("Matching Run Summary", 14, bold=True))
        self.log_box = QTextEdit(); self.log_box.setReadOnly(True)
        self.log_box.setPlaceholderText("Run matching to see the guided summary…")
        lay.addWidget(self.log_box)
        return w

    # ══════════════════════════════════════════════════════
    #  ACTIONS
    # ══════════════════════════════════════════════════════
    def _browse(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel Input", "",
            "Excel Files (*.xlsx *.xls);;All Files (*)")
        if path: self._load(path)

    def _load(self, path):
        try:
            courses, tas, caps, grade, ta_u, cu, k = read_excel(path)
            self._data = dict(
                courses=courses, tas=tas, capacities=caps,
                grade=grade, ta_utility=ta_u,
                course_utility=cu, k=k, path=path,
                matching=None, report=None, nc=None, nt=None, meta={})
            analysis = analyze_matching_data(courses, tas, caps, grade, ta_u, cu)
            self._data['analysis'] = analysis

            self.file_lbl.setText(f"✅  {os.path.basename(path)}")
            self.file_lbl.setStyleSheet(
                f"color:{GREEN};font-size:13px;background:transparent;")

            self._upd_sc("Courses",     f"{len(courses)}")
            self._upd_sc("TAs",         f"{len(tas)}")
            self._upd_sc("Grade",       f"{len(tas)}×{len(courses)}")
            self._upd_sc("TA Utility",  f"{len(tas)}×{len(courses)}")
            self._upd_sc("Course Util", f"{len(courses)}×{len(tas)}")
            self._upd_sc("k",           f"k={k}")

            self._show_previews(courses, tas, caps, grade, ta_u, cu)
            self.smart_panel.update_analysis(analysis)
            self.strategy_status_lbl.setText(analysis['recommendation_text'])
            self.subtitle_lbl.setText(
                f"Recommended: {next(s['title'] for s in SMART_STRATEGIES if s['id'] == analysis['recommended_strategy'])}"
            )
            self.btn_run.setEnabled(True)
        except Exception as e:
            QMessageBox.critical(self, "Error reading file", str(e))

    def _show_previews(self, courses, tas, caps, grade, ta_u, cu):
        self.prev_tabs.clear(); self.prev_tabs.setVisible(True)

        t = make_table(["Course", "Capacity"])
        for c in courses: add_row(t, [c, caps[c]])
        self.prev_tabs.addTab(t, "Courses")

        t2 = make_table(["TA Name"])
        for ta in tas: add_row(t2, [ta])
        self.prev_tabs.addTab(t2, "TAs")

        for title, mat, rows, cols, accent in [
            ("Grade",       grade, tas,     courses, "#ffd700"),
            ("TA Utility",  ta_u,  tas,     courses, ACCENT2),
            ("Course Util", cu,    courses, tas,     ACCENT),
        ]:
            tw = QTableWidget(len(rows), len(cols))
            tw.setHorizontalHeaderLabels(cols)
            tw.setVerticalHeaderLabels(rows)
            tw.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            tw.setEditTriggers(QTableWidget.NoEditTriggers)
            for i, r in enumerate(rows):
                for j, c in enumerate(cols):
                    val  = mat.get((r, c), 0)
                    item = QTableWidgetItem(str(val))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setForeground(QColor(accent if val != 0 else SUBTEXT))
                    tw.setItem(i, j, item)
            self.prev_tabs.addTab(tw, title)

    def _run(self):
        if not self._data: return
        self.btn_run.setEnabled(False)
        self.btn_run.setText("  ⏳  Running…")
        d = self._data

        analysis = d.get('analysis') or analyze_matching_data(
            d['courses'], d['tas'], d['capacities'],
            d['grade'], d['ta_utility'], d['course_utility'])
        d['analysis'] = analysis

        strategy_id = self.smart_panel.selected_strategy()
        custom_backend = self.smart_panel.selected_custom_backend()
        result, log_text = execute_strategy(d, strategy_id, analysis, custom_backend)

        matching = result['matching'] if result else None
        report = result['report'] if result else None
        d['matching'] = matching
        d['report'] = report
        d['nc'] = result['nc'] if result else analysis['nc']
        d['nt'] = result['nt'] if result else analysis['nt']
        d['meta'] = result['meta'] if result else {}
        d['selected_strategy'] = strategy_id
        d['engine_label'] = result['backend_label'] if result else "No engine selected"

        # Log tab
        self.log_box.setPlainText(log_text)

        # Verdict banner
        if matching:
            self.verdict.setText(f"✅  Matching complete using {d['engine_label']}")
            self.verdict.setStyleSheet(
                f"background:#1a3a2a;color:{GREEN};border-radius:10px;padding:14px;"
                f"font-size:16px;font-weight:700;border:2px solid {GREEN};")
        else:
            self.verdict.setText("❌  No valid matching could be completed with the current data.")
            self.verdict.setStyleSheet(
                f"background:#3a1a1a;color:{DANGER};border-radius:10px;padding:14px;"
                f"font-size:16px;font-weight:700;border:2px solid {DANGER};")

        # Matching table
        self.match_tbl.setRowCount(0)
        if matching:
            for ta, c in sorted(matching.items()):
                add_row(self.match_tbl, [ta, c], [ACCENT2, ACCENT])
        for ta in d['tas']:
            if not matching or ta not in matching:
                add_row(self.match_tbl,
"""  """                        [ta, "— unassigned —"], [SUBTEXT, SUBTEXT])

        # Checks table
        self.chk_tbl.setRowCount(0)
        if report:
            for name, info in report.items():
                ok = info['passed']
                add_row(self.chk_tbl,
                        [name, "✅ PASS" if ok else "❌ FAIL", info['message']],
                        [None, GREEN if ok else DANGER, SUBTEXT])

        # Course summary
        self.course_tbl.setRowCount(0)
        if matching and report:
            assigned  = defaultdict(list)
            for ta, c in matching.items(): assigned[c].append(ta)
            avg_utils = report.get('Satisfaction', {}).get('avg_utils', {})
            kv = d['k']
            for c in d['courses']:
                tl  = sorted(assigned.get(c, []))
                avg = avg_utils.get(c, 0)
                ok  = avg >= kv
                add_row(self.course_tbl,
                        [c, d['capacities'][c],
                         ", ".join(tl) or "—",
                         f"{avg:.2f}", "YES" if ok else "NO"],
                        [ACCENT, None, ACCENT2,
                         GREEN if ok else DANGER,
                         GREEN if ok else DANGER])

        self.btn_run.setEnabled(True)
        self.btn_run.setText("  Run Matching  ")
        self.btn_exp.setEnabled(bool(matching))

        self._refresh_graph()
        self.tabs.setCurrentIndex(2)

    def _refresh_graph(self):
        d = self._data
        if not d or not d.get('nc'): return

        wt_map  = ["TA Utility", "Course Utility", "Grade", "Both u & v"]
        wt_mode = wt_map[self.cmb_weight.currentIndex()]
        after   = self.cmb_view.currentIndex() == 0
        match   = d.get('matching') if after else {}

        self.canvas.draw_graph(
            d['courses'], d['tas'], d['capacities'],
            d['nc'], match or {},
            d.get('meta', {}),
            self.chk_weights.isChecked(),
            d['ta_utility'], d['course_utility'],
            d['grade'], wt_mode)

        case_names = {
            'case1': 'Auto Fill',          'case2_1': 'Tree Resolution',
            'case2_2': 'Cycle Resolution', 'case2_3': 'Multi-cycle Review',
            'thm5':  'Preference-First',   'thm6':    'Quick Single-Choice',
            'thm7':  'Deep Search',        'thm9':    'Exhaustive Search',
            'thm12': 'Exchange-Based',     'thm13':   'Full Preference',
        }
        parts = []
        for idx, m in d.get('meta', {}).items():
            cn = case_names.get(m['case'], m['case'])
            parts.append(f"Comp {idx+1}: {cn}  "
                         f"({len(m['courses'])} courses, {len(m['tas'])} TAs)")
        self.graph_info.setText(
            "   |   ".join(parts) if parts else "No graph data yet.")

    # ── Export ───────────────────────────────────────────
    def _export(self):
        if not self._data.get('matching'): return
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Result", "mefe_result.xlsx",
            "Excel Files (*.xlsx)")
        if not path: return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb  = openpyxl.Workbook()
            ws  = wb.active; ws.title = "Matching"
            hf  = Font(bold=True, color="FFFFFF")
            hb  = PatternFill("solid", fgColor="2F5496")
            gf  = PatternFill("solid", fgColor="C6EFCE")
            rf  = PatternFill("solid", fgColor="FFC7CE")
            ctr = Alignment(horizontal="center")
            for col, h in enumerate(["TA", "Assigned Course"], 1):
                c = ws.cell(1, col, h); c.font = hf; c.fill = hb; c.alignment = ctr
            for r, (ta, c) in enumerate(
                    sorted(self._data['matching'].items()), 2):
                ws.cell(r, 1, ta); ws.cell(r, 2, c)
            ws2 = wb.create_sheet("Validity")
            for col, h in enumerate(["Check", "Result", "Message"], 1):
                c = ws2.cell(1, col, h); c.font = hf; c.fill = hb
            for r, (name, info) in enumerate(
                    (self._data.get('report') or {}).items(), 2):
                ws2.cell(r, 1, name)
                res = ws2.cell(r, 2, "PASS" if info['passed'] else "FAIL")
                res.fill = gf if info['passed'] else rf
                ws2.cell(r, 3, info['message'])
            wb.save(path)
            QMessageBox.information(self, "Saved", f"Saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))

    # ── Sample file ──────────────────────────────────────
    def _make_sample(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Sample", "mefe_sample.xlsx",
            "Excel Files (*.xlsx)")
        if not path: return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb  = openpyxl.Workbook()
            hf  = Font(bold=True, color='FFFFFF')
            hb  = PatternFill('solid', fgColor='2F5496')
            ctr = Alignment(horizontal='center')
            def hdr(c): c.font = hf; c.fill = hb; c.alignment = ctr
            ws = wb.active; ws.title = 'Courses'
            hdr(ws.cell(1,1,'Course Name')); hdr(ws.cell(1,2,'Capacity'))
            for i,(n,c) in enumerate([('x1',2),('x2',1)],2):
                ws.cell(i,1,n); ws.cell(i,2,c)
            ws2 = wb.create_sheet('TAs'); hdr(ws2.cell(1,1,'TA Name'))
            for i,t in enumerate(['t1','t2','t3','t4'],2): ws2.cell(i,1,t)
            ws3 = wb.create_sheet('Grade_g'); ws3.cell(1,1,'')
            for j,c in enumerate(['x1','x2'],2): hdr(ws3.cell(1,j,c))
            for i,(ta,g1,g2) in enumerate(
                    [('t1',3,0),('t2',8,0),('t3',5,4),('t4',0,9)],2):
                ws3.cell(i,1,ta); ws3.cell(i,2,g1); ws3.cell(i,3,g2)
            ws4 = wb.create_sheet('TA_Utility_u'); ws4.cell(1,1,'')
            for j,c in enumerate(['x1','x2'],2): hdr(ws4.cell(1,j,c))
            for i,(ta,u1,u2) in enumerate(
                    [('t1',0,0),('t2',3,0),('t3',1,3),('t4',0,2)],2):
                ws4.cell(i,1,ta); ws4.cell(i,2,u1); ws4.cell(i,3,u2)
            ws5 = wb.create_sheet('Course_Utility_v'); ws5.cell(1,1,'')
            for j,ta in enumerate(['t1','t2','t3','t4'],2):
                hdr(ws5.cell(1,j,ta))
            for i,(c,vals) in enumerate(
                    [('x1',[0,3,2,0]),('x2',[0,0,5,4])],2):
                ws5.cell(i,1,c)
                for j,v in enumerate(vals,2): ws5.cell(i,j,v)
            ws6 = wb.create_sheet('Threshold_k')
            hdr(ws6.cell(1,1,'k')); ws6.cell(2,1,2)
            wb.save(path)
            QMessageBox.information(
                self, "Sample Created",
                f"Saved to:\n{path}\n\nNow upload it using Browse File.\n"
                f"This sample is optimized for the balanced matching engine.")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))


# ═══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    win = MEFEApp()
    win.show()
    sys.exit(app.exec_())
