import sys, os
from collections import defaultdict, deque

import pandas as pd
import networkx as nx
import matplotlib
matplotlib.use('Qt5Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QTableWidget, QTableWidgetItem,
    QTabWidget, QTextEdit, QFrame, QHeaderView, QMessageBox, QComboBox,
    QCheckBox
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor


# ── Palette ──────────────────────────────────────────────────
BG      = "#0f1117";  PANEL  = "#1a1d2e";  CARD   = "#21253a"
ACCENT  = "#6c63ff";  ACCENT2= "#00d4aa";  DANGER = "#ff4757"
TEXT    = "#e8e8f0";  SUBTEXT= "#8888aa";  BORDER = "#2e3150"
GREEN   = "#2ed573"

# ── Graph colours ─────────────────────────────────────────────
COL_COURSE     = "#6c63ff"
COL_TA         = "#00d4aa"
COL_EDGE       = "#555577"
COL_MATCHED    = "#ffd700"
COL_UNASSIGNED = "#ff4757"
COL_CYCLE      = "#ff9f43"
COL_FORCED     = "#54a0ff"


# ═══════════════════════════════════════════════════════════════
#  ALGORITHM
# ═══════════════════════════════════════════════════════════════

def read_excel(filepath):
    df_c = pd.read_excel(filepath, sheet_name=0)
    courses    = list(df_c.iloc[:, 0].astype(str))
    capacities = {courses[i]: int(df_c.iloc[i, 1]) for i in range(len(courses))}
    df_t = pd.read_excel(filepath, sheet_name=1)
    tas  = list(df_t.iloc[:, 0].astype(str))

    def load_matrix(idx):
        df = pd.read_excel(filepath, sheet_name=idx, index_col=0)
        df.index   = [str(i) for i in df.index]
        df.columns = [str(c) for c in df.columns]
        return {(r, c): float(df.loc[r, c]) for r in df.index for c in df.columns}

    grade          = load_matrix(2)
    ta_utility     = load_matrix(3)
    course_utility = load_matrix(4)
    df_k = pd.read_excel(filepath, sheet_name=5)
    k    = float(df_k.iloc[0, 0])
    return courses, tas, capacities, grade, ta_utility, course_utility, k


def build_graph(courses, tas, ta_utility):
    nc, nt = defaultdict(set), defaultdict(set)
    for (ta, c), u in ta_utility.items():
        if u != 0:
            nc[c].add(ta); nt[ta].add(c)
    return dict(nc), dict(nt)


def get_components(courses, tas, nc, nt):
    vc, vt = set(), set()
    comps  = []
    def bfs(start):
        cc, ct = set(), set()
        q = deque([('c', start)])
        while q:
            kind, node = q.popleft()
            if kind == 'c':
                if node in vc: continue
                vc.add(node); cc.add(node)
                for t in nc.get(node, []):
                    if t not in vt: q.append(('t', t))
            else:
                if node in vt: continue
                vt.add(node); ct.add(node)
                for c in nt.get(node, []):
                    if c not in vc: q.append(('c', c))
        return cc, ct
    for c in courses:
        if c not in vc:
            cc, ct = bfs(c)
            if cc: comps.append((cc, ct))
    return comps


def find_cycle(cc, ct, nc, nt):
    visited, parent = {}, {}
    def nbrs(kind, node):
        return [('t', t) for t in nc.get(node, [])] if kind == 'c' \
               else [('c', c) for c in nt.get(node, [])]
    def dfs(kind, node, pk, pn):
        key = (kind, node)
        visited[key] = True
        for nk, nn in nbrs(kind, node):
            nkey = (nk, nn)
            if nkey == (pk, pn): continue
            if nkey in visited:
                path = [nkey, key]
                cur  = parent.get(key)
                while cur and cur != nkey:
                    path.append(cur); cur = parent.get(cur)
                return path
            parent[nkey] = key
            r = dfs(nk, nn, kind, node)
            if r: return r
        return None
    for kind, node in [('c', c) for c in cc] + [('t', t) for t in ct]:
        key = (kind, node)
        if key not in visited:
            parent[key] = None
            r = dfs(kind, node, None, None)
            if r:
                return [n for (k, n) in r if k == 'c'], \
                       [n for (k, n) in r if k == 't']
    return [], []


def extended_matching(comp_courses, nc, matching, mt, mc):
    changed = True
    while changed:
        changed = False
        for course in comp_courses:
            if course in mc: continue
            nbrs    = nc.get(course, set())
            overlap = nbrs & mt
            if overlap:
                for ta in nbrs - mt:
                    matching[ta] = course; mt.add(ta)
                mc.add(course); changed = True


def solve(courses, tas, capacities, grade, ta_utility, course_utility, k):
    log  = []
    nc, nt = build_graph(courses, tas, ta_utility)
    meta   = {}

    for c in courses:
        d, cap = len(nc.get(c, set())), capacities[c]
        log.append(f"Course '{c}': degree={d}, capacity={cap}, diff={d-cap}")
        if d - cap > 1:
            log.append("✗ Theorem 4 does NOT apply (diff > 1)")
            return None, None, log, nc, nt, meta

    comps = get_components(courses, tas, nc, nt)
    log.append(f"\nConnected components: {len(comps)}")
    full_matching = {}

    for idx, (cc, ct) in enumerate(comps):
        log.append(f"\nComponent {idx+1}: courses={sorted(cc)}, TAs={sorted(ct)}")
        E = sum(len(nc.get(c, set())) for c in cc)
        V = len(cc) + len(ct)

        has_zero = any(len(nc.get(c, set())) - capacities[c] == 0 for c in cc)
        has_neg  = any(len(nc.get(c, set())) - capacities[c] <  0 for c in cc)

        if has_neg:
            log.append("  ✗ Not enough TAs")
            return None, None, log, nc, nt, meta

        case = ('case1'   if has_zero else
                'case2_1' if E == V-1 else
                'case2_2' if E == V   else 'case2_3')

        log.append(f"  Case: {case}")
        meta[idx] = {'case': case, 'courses': cc, 'tas': ct,
                     'cycle_courses': [], 'cycle_tas': [],
                     'unassigned_ta': None, 'forced_courses': set()}

        if case == 'case2_3':
            log.append("  ✗ Multiple cycles")
            return None, None, log, nc, nt, meta

        comp_match = {}

        if case == 'case1':
            matching, mt, mc = {}, set(), set()
            forced = set()
            for c in cc:
                if len(nc.get(c, set())) == capacities[c]:
                    for ta in nc.get(c, set()):
                        matching[ta] = c; mt.add(ta)
                    mc.add(c); forced.add(c)
                    log.append(f"  Forced: {c} ← {sorted(nc.get(c, set()))}")
            extended_matching(cc, nc, matching, mt, mc)
            comp_match = matching
            meta[idx]['forced_courses'] = forced

        elif case == 'case2_1':
            found = False
            for unassigned in ct:
                matching, mt, mc = {}, {unassigned}, set()
                extended_matching(cc, nc, matching, mt, mc)
                matching.pop(unassigned, None)
                if len(mc) == len(cc):
                    log.append(f"  TA '{unassigned}' unassigned → valid")
                    comp_match = matching
                    meta[idx]['unassigned_ta'] = unassigned
                    found = True; break
            if not found:
                log.append("  ✗ No valid tree matching")
                return None, None, log, nc, nt, meta

        elif case == 'case2_2':
            cyc_c, cyc_t = find_cycle(cc, ct, nc, nt)
            meta[idx]['cycle_courses'] = cyc_c
            meta[idx]['cycle_tas']     = cyc_t
            log.append(f"  Cycle: {cyc_c}")
            x1 = cyc_c[0] if cyc_c else list(cc)[0]
            nbrs_x1 = list(nc.get(x1, set()))
            found = False
            for excl in nbrs_x1[:2]:
                matching, mt, mc = {}, {excl}, set()
                for ta in nbrs_x1:
                    if ta != excl:
                        matching[ta] = x1; mt.add(ta)
                mc.add(x1)
                extended_matching(cc, nc, matching, mt, mc)
                if len(mc) == len(cc):
                    log.append(f"  Excluded '{excl}' from '{x1}' → valid")
                    comp_match = matching; found = True; break
            if not found:
                log.append("  ✗ No valid cycle matching")
                return None, None, log, nc, nt, meta

        full_matching.update(comp_match)

    report = validate(full_matching, courses, tas, capacities,
                      nc, course_utility, ta_utility, grade, k)
    log.append("\n── VALIDATION ──")
    for key, info in report.items():
        log.append(f"  {'✅' if info['passed'] else '❌'} {key}: {info['message']}")
    ok = all(v['passed'] for v in report.values())
    log.append(f"\n{'✅ VALID MEFE MATCHING FOUND' if ok else '❌ NO VALID MEFE MATCHING'}")
    return (full_matching if ok else None), report, log, nc, nt, meta


def validate(matching, courses, tas, capacities, nc,
             course_utility, ta_utility, grade, k):
    report   = {}
    assigned = defaultdict(list)
    for ta, c in matching.items(): assigned[c].append(ta)

    ok = True; msg = "OK"
    for c in courses:
        if len(assigned.get(c, [])) != capacities[c]:
            ok  = False
            msg = f"'{c}' needs {capacities[c]}, got {len(assigned.get(c,[]))}"
            break
    report['Feasibility'] = {'passed': ok, 'message': msg}

    ok = True; msg = "OK"; avg_utils = {}
    for c in courses:
        tl  = assigned.get(c, [])
        cap = capacities[c]
        if not cap: continue
        avg = sum(course_utility.get((c, t), 0) for t in tl) / cap
        avg_utils[c] = round(avg, 2)
        if avg < k:
            ok = False; msg = f"'{c}' AvgUtil={avg:.2f}<k={k}"; break
    report['Satisfaction'] = {'passed': ok, 'message': msg, 'avg_utils': avg_utils}

    ok = True; msg = "OK"
    for ti in tas:
        ci = matching.get(ti)
        for tj in tas:
            if ti == tj: continue
            cj = matching.get(tj)
            if not cj: continue
            gi_j = grade.get((ti, cj), 0)
            gj_j = grade.get((tj, cj), 0)
            ui_j = ta_utility.get((ti, cj), 0)
            ui_i = ta_utility.get((ti, ci), 0) if ci else 0
            if gi_j >= gj_j and ui_j > ui_i:
                ok = False; msg = f"'{ti}' envies '{tj}' (course '{cj}')"; break
        if not ok: break
    report['Envy-Freeness'] = {'passed': ok, 'message': msg}
    return report


# ═══════════════════════════════════════════════════════════════
#  GRAPH CANVAS
# ═══════════════════════════════════════════════════════════════

class GraphCanvas(FigureCanvas):
    def __init__(self, parent=None):
        self.fig = Figure(facecolor='#0f1117', tight_layout=False)
        super().__init__(self.fig)
        self.setParent(parent)
        self._placeholder()

    def _placeholder(self):
        self.fig.clear()
        ax = self.fig.add_subplot(111)
        ax.set_facecolor('#0f1117')
        ax.text(0.5, 0.5,
                'Upload Excel  →  Run Algorithm\nto visualize the Bipartite Graph',
                ha='center', va='center', fontsize=15,
                color='#8888aa', transform=ax.transAxes,
                fontfamily='monospace')
        ax.axis('off')
        self.draw()

    # ── Public draw entry ─────────────────────────────────────
    def draw_graph(self, courses, tas, capacities, nc,
                   matching, meta, show_weights,
                   ta_utility, course_utility, grade, weight_mode):
        self.fig.clear()
        comps = list(meta.values())
        if not comps:
            self._placeholder(); return

        n      = len(comps)
        ncols  = min(n, 2)
        nrows  = (n + 1) // 2
        axes   = self.fig.subplots(nrows, ncols, squeeze=False)

        for idx, m in enumerate(comps):
            ax = axes[idx // ncols][idx % ncols]
            ax.set_facecolor('#13162a')
            self._draw_comp(ax, m, nc, capacities, matching,
                            show_weights, ta_utility, course_utility,
                            grade, weight_mode, idx + 1)

        # hide unused axes
        for idx in range(n, nrows * ncols):
            axes[idx // ncols][idx % ncols].set_visible(False)

        self._legend()
        self.fig.subplots_adjust(hspace=0.4, wspace=0.35,
                                 bottom=0.12, top=0.95)
        self.draw()

    # ── Draw one component ────────────────────────────────────
    def _draw_comp(self, ax, m, nc, capacities, matching,
                   show_weights, ta_u, cu, grade, wmode, comp_idx):
        cc     = m['courses']
        ct     = m['tas']
        case   = m['case']
        cyc_c  = set(m.get('cycle_courses', []))
        cyc_t  = set(m.get('cycle_tas',     []))
        forced = m.get('forced_courses', set())
        unass  = m.get('unassigned_ta')

        cc_list = sorted(cc)
        ct_list = sorted(ct)

        # ── Build networkx graph ──
        G = nx.Graph()
        for c in cc_list: G.add_node(c, kind='course')
        for t in ct_list: G.add_node(t, kind='ta')
        for c in cc_list:
            for t in nc.get(c, []):
                if t in ct: G.add_edge(c, t)

        # ── Positions: courses left, TAs right ──
        pos = {}
        for i, c in enumerate(cc_list):
            y = 0 if len(cc_list) == 1 else 1 - i * 2.0 / (len(cc_list) - 1)
            pos[c] = (-1.0, y)
        for i, t in enumerate(ct_list):
            y = 0 if len(ct_list) == 1 else 1 - i * 2.0 / (len(ct_list) - 1)
            pos[t] = (1.0, y)

        # ── Node colours / sizes ──
        node_col  = []
        node_size = []
        node_ec   = []
        for node in G.nodes():
            kind = G.nodes[node]['kind']
            if kind == 'course':
                if node in forced:       col = COL_FORCED
                elif node in cyc_c:      col = COL_CYCLE
                else:                    col = COL_COURSE
                node_size.append(2000)
            else:
                if node == unass:        col = COL_UNASSIGNED
                elif node in cyc_t:      col = COL_CYCLE
                else:                    col = COL_TA
                node_size.append(1600)
            node_col.append(col)
            node_ec.append('#ffffff')

        # ── Matched edge set ──
        matched_pairs = set()
        if matching:
            for ta, c in matching.items():
                if ta in ct and c in cc:
                    matched_pairs.add((c, ta))
                    matched_pairs.add((ta, c))

        # ── Separate matched / unmatched edges ──
        solid, dashed = [], []
        for u, v in G.edges():
            if (u, v) in matched_pairs or (v, u) in matched_pairs:
                solid.append((u, v))
            else:
                dashed.append((u, v))

        # ── Draw nodes ──
        nx.draw_networkx_nodes(G, pos, ax=ax,
                               node_color=node_col,
                               node_size=node_size,
                               edgecolors=node_ec,
                               linewidths=1.8)

        # ── Draw edges ──
        if solid:
            nx.draw_networkx_edges(G, pos, edgelist=solid, ax=ax,
                                   edge_color=COL_MATCHED,
                                   width=3.5, style='solid', alpha=0.95)
        if dashed:
            nx.draw_networkx_edges(G, pos, edgelist=dashed, ax=ax,
                                   edge_color=COL_EDGE,
                                   width=1.2, style='dashed', alpha=0.55)

        # ── Node labels ──
        nx.draw_networkx_labels(G, pos, ax=ax,
                                font_size=9, font_color='white',
                                font_weight='bold')

        # ── Capacity annotations ──
        for c in cc_list:
            x, y = pos[c]
            ax.annotate(f"cap={capacities.get(c,'?')}",
                        xy=(x, y), xytext=(x - 0.38, y),
                        fontsize=7, color='#aaaacc',
                        ha='right', va='center',
                        arrowprops=None)

        # ── Edge weight labels ──
        if show_weights:
            elabels = {}
            for u, v in G.edges():
                c_n = u if G.nodes[u]['kind'] == 'course' else v
                t_n = v if G.nodes[u]['kind'] == 'course' else u
                if wmode == 'TA Utility':
                    elabels[(u, v)] = f"u={ta_u.get((t_n,c_n),0)}"
                elif wmode == 'Course Utility':
                    elabels[(u, v)] = f"v={cu.get((c_n,t_n),0)}"
                elif wmode == 'Grade':
                    elabels[(u, v)] = f"g={grade.get((t_n,c_n),0)}"
                else:
                    elabels[(u, v)] = (f"u={ta_u.get((t_n,c_n),0)}\n"
                                       f"v={cu.get((c_n,t_n),0)}")

            nx.draw_networkx_edge_labels(
                G, pos, elabels, ax=ax,
                font_size=7, font_color='#ffdd88',
                label_pos=0.35,
                bbox=dict(boxstyle='round,pad=0.2',
                          facecolor='#1a1d2e',
                          edgecolor='none', alpha=0.85))

        # ── Column headers ──
        ylim = ax.get_ylim()
        top  = ylim[1] if ylim[1] != ylim[0] else 1.3
        ax.text(-1.0, top + 0.08, 'COURSES', ha='center', va='bottom',
                fontsize=9, color=COL_COURSE, fontweight='bold',
                transform=ax.transData)
        ax.text( 1.0, top + 0.08, 'TAs', ha='center', va='bottom',
                fontsize=9, color=COL_TA, fontweight='bold',
                transform=ax.transData)

        case_label = {
            'case1':   '⚡ Case 1 — Forced',
            'case2_1': '🌲 Case 2.1 — Tree',
            'case2_2': '🔄 Case 2.2 — One Cycle',
            'case2_3': '❌ Case 2.3 — Multi-cycle',
        }.get(case, case)

        ax.set_title(f"Component {comp_idx}  |  {case_label}",
                     color='white', fontsize=11,
                     fontweight='bold', pad=14)
        ax.set_xlim(-1.85, 1.85)
        ax.axis('off')

    # ── Bottom legend ─────────────────────────────────────────
    def _legend(self):
        patches = [
            mpatches.Patch(color=COL_COURSE,     label='Course'),
            mpatches.Patch(color=COL_TA,         label='TA'),
            mpatches.Patch(color=COL_FORCED,     label='Forced (Case 1)'),
            mpatches.Patch(color=COL_CYCLE,      label='Cycle node'),
            mpatches.Patch(color=COL_UNASSIGNED, label='Unassigned TA'),
            mpatches.Patch(color=COL_MATCHED,    label='Matched edge'),
            mpatches.Patch(color=COL_EDGE,       label='Unmatched edge'),
        ]
        self.fig.legend(handles=patches,
                        loc='lower center', ncol=4,
                        frameon=True,
                        facecolor='#21253a', edgecolor=BORDER,
                        labelcolor='white', fontsize=9,
                        bbox_to_anchor=(0.5, 0.0))


# ═══════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════

def lbl(text, size=13, bold=False, color=TEXT):
    w = QLabel(text)
    w.setStyleSheet(f"color:{color};font-size:{size}px;"
                    f"{'font-weight:bold;' if bold else ''}background:transparent;")
    return w


def make_btn(text, bg=ACCENT, hover=None):
    hover = hover or bg
    b = QPushButton(text)
    b.setCursor(Qt.PointingHandCursor)
    b.setStyleSheet(f"""
        QPushButton{{background:{bg};color:#fff;border:none;
            border-radius:8px;padding:10px 22px;
            font-size:13px;font-weight:600;}}
        QPushButton:hover{{background:{hover};}}
        QPushButton:disabled{{background:#333;color:#666;}}""")
    return b


def card_frame():
    f = QFrame()
    f.setStyleSheet(
        f"background:{CARD};border-radius:12px;border:1px solid {BORDER};")
    return f


def make_table(headers):
    t = QTableWidget(0, len(headers))
    t.setHorizontalHeaderLabels(headers)
    t.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    t.verticalHeader().setVisible(False)
    t.setEditTriggers(QTableWidget.NoEditTriggers)
    t.setAlternatingRowColors(True)
    t.setStyleSheet(
        t.styleSheet() +
        "QTableWidget{alternate-background-color:#1e2235;}")
    return t


def add_row(tbl, vals, colors=None):
    r = tbl.rowCount(); tbl.insertRow(r)
    for col, val in enumerate(vals):
        item = QTableWidgetItem(str(val))
        item.setTextAlignment(Qt.AlignCenter)
        if colors and col < len(colors) and colors[col]:
            item.setForeground(QColor(colors[col]))
        tbl.setItem(r, col, item)


# ═══════════════════════════════════════════════════════════════
#  MAIN WINDOW
# ═══════════════════════════════════════════════════════════════

class MEFEApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("MEFE Matching — Theorem 4  |  Graph Visualization")
        self.setMinimumSize(1200, 800)
        self.resize(1350, 900)
        self._data = {}
        self._apply_style()
        self._build_ui()

    def _apply_style(self):
        self.setStyleSheet(f"""
            QMainWindow,QWidget{{
                background:{BG};color:{TEXT};
                font-family:'Segoe UI',sans-serif;}}
            QTabWidget::pane{{
                border:1px solid {BORDER};border-radius:8px;
                background:{PANEL};}}
            QTabBar::tab{{
                background:{CARD};color:{SUBTEXT};padding:10px 20px;
                border-top-left-radius:8px;border-top-right-radius:8px;
                font-size:13px;}}
            QTabBar::tab:selected{{
                background:{ACCENT};color:#fff;font-weight:bold;}}
            QTableWidget{{
                background:{PANEL};color:{TEXT};
                gridline-color:{BORDER};border:none;
                border-radius:8px;font-size:12px;}}
            QTableWidget::item:selected{{background:{ACCENT};}}
            QHeaderView::section{{
                background:{CARD};color:{SUBTEXT};border:none;
                padding:6px;font-weight:bold;font-size:12px;}}
            QTextEdit{{
                background:{PANEL};color:{TEXT};
                border:1px solid {BORDER};border-radius:8px;
                font-family:'Consolas',monospace;
                font-size:12px;padding:8px;}}
            QComboBox{{
                background:{CARD};color:{TEXT};
                border:1px solid {BORDER};border-radius:6px;
                padding:6px 12px;font-size:12px;min-width:160px;}}
            QComboBox::drop-down{{border:none;}}
            QComboBox QAbstractItemView{{
                background:{CARD};color:{TEXT};
                selection-background-color:{ACCENT};}}
            QCheckBox{{color:{TEXT};font-size:12px;spacing:6px;}}
            QCheckBox::indicator{{
                width:16px;height:16px;border-radius:4px;
                border:1px solid {BORDER};background:{PANEL};}}
            QCheckBox::indicator:checked{{
                background:{ACCENT};border-color:{ACCENT};}}
            QScrollBar:vertical{{
                background:{PANEL};width:8px;border-radius:4px;}}
            QScrollBar::handle:vertical{{
                background:{ACCENT};border-radius:4px;}}
        """)

    # ── scaffold ─────────────────────────────────────────────
    def _build_ui(self):
        root = QWidget(); self.setCentralWidget(root)
        main = QVBoxLayout(root)
        main.setContentsMargins(20, 20, 20, 20); main.setSpacing(14)

        hdr = QHBoxLayout()
        hdr.addWidget(lbl("MEFE Matching", 22, bold=True, color=ACCENT))
        hdr.addWidget(lbl("Theorem 4  •  Bipartite Graph Visualization",
                          13, color=SUBTEXT))
        hdr.addStretch()
        main.addLayout(hdr)

        self.tabs = QTabWidget()
        self.tabs.addTab(self._tab_upload(),  "📂  Upload")
        self.tabs.addTab(self._tab_results(), "⚙️  Results")
        self.tabs.addTab(self._tab_graph(),   "🕸️  Graph")
        self.tabs.addTab(self._tab_log(),     "📋  Log")
        main.addWidget(self.tabs)

    # ══════════════════════════════════════════════════════
    #  TAB 1 — UPLOAD
    # ══════════════════════════════════════════════════════
    def _tab_upload(self):
        w = QWidget(); lay = QVBoxLayout(w); lay.setSpacing(14)

        zone = card_frame()
        zl   = QVBoxLayout(zone)
        zl.setAlignment(Qt.AlignCenter)
        zl.setContentsMargins(40, 35, 40, 35)
        zl.setSpacing(10)

        ic = lbl("📂", 46, color=ACCENT); ic.setAlignment(Qt.AlignCenter)
        zl.addWidget(ic)
        zl.addWidget(lbl("Upload Excel Input File", 16, bold=True))
        zl.addWidget(lbl(
            "6 sheets: Courses · TAs · Grade_g · "
            "TA_Utility_u · Course_Utility_v · Threshold_k",
            12, color=SUBTEXT))

        br = QHBoxLayout()
        self.btn_browse = make_btn("  Browse File  ", ACCENT)
        self.btn_sample = make_btn("  Download Sample  ", ACCENT2)
        self.btn_browse.clicked.connect(self._browse)
        self.btn_sample.clicked.connect(self._make_sample)
        br.addStretch()
        br.addWidget(self.btn_browse); br.addWidget(self.btn_sample)
        br.addStretch()
        zl.addLayout(br)

        self.file_lbl = lbl("No file selected", 12, color=SUBTEXT)
        self.file_lbl.setAlignment(Qt.AlignCenter)
        zl.addWidget(self.file_lbl)
        lay.addWidget(zone)

        # status cards
        self._sc = {}
        sc_row   = QHBoxLayout()
        for name in ["Courses", "TAs", "Grade", "TA Utility",
                     "Course Util", "k"]:
            f  = card_frame(); f.setFixedHeight(78)
            vl = QVBoxLayout(f); vl.setAlignment(Qt.AlignCenter)
            t  = lbl(name, 10, color=SUBTEXT); t.setAlignment(Qt.AlignCenter)
            v  = lbl("—", 15, bold=True, color=ACCENT2)
            v.setAlignment(Qt.AlignCenter); v.setObjectName("val")
            vl.addWidget(t); vl.addWidget(v)
            self._sc[name] = f; sc_row.addWidget(f)
        lay.addLayout(sc_row)

        self.prev_tabs = QTabWidget(); self.prev_tabs.setVisible(False)
        lay.addWidget(self.prev_tabs)
        return w

    def _upd_sc(self, name, val, ok=True):
        v = self._sc[name].findChild(QLabel, "val")
        if v:
            v.setText(str(val))
            c = GREEN if ok else DANGER
            v.setStyleSheet(
                f"color:{c};font-size:15px;"
                f"font-weight:bold;background:transparent;")

    # ══════════════════════════════════════════════════════
    #  TAB 2 — RESULTS
    # ══════════════════════════════════════════════════════
    def _tab_results(self):
        w = QWidget(); lay = QVBoxLayout(w); lay.setSpacing(14)

        run_row = QHBoxLayout()
        self.btn_run = make_btn("  ▶  Run Algorithm  ", ACCENT)
        self.btn_run.setFixedHeight(46)
        self.btn_run.setEnabled(False)
        self.btn_run.clicked.connect(self._run)
        run_row.addStretch(); run_row.addWidget(self.btn_run); run_row.addStretch()
        lay.addLayout(run_row)

        self.verdict = QLabel("Upload an Excel file first, then click Run")
        self.verdict.setAlignment(Qt.AlignCenter)
        self.verdict.setStyleSheet(
            f"background:{CARD};color:{SUBTEXT};"
            f"border-radius:10px;padding:14px;"
            f"font-size:15px;font-weight:600;"
            f"border:1px solid {BORDER};")
        lay.addWidget(self.verdict)

        panels = QHBoxLayout(); panels.setSpacing(14)

        lc = QVBoxLayout()
        lc.addWidget(lbl("Matching Result", 13, bold=True))
        self.match_tbl = make_table(["TA", "Assigned Course"])
        lc.addWidget(self.match_tbl); panels.addLayout(lc, 1)

        rc = QVBoxLayout()
        rc.addWidget(lbl("Validity Checks", 13, bold=True))
        self.chk_tbl = make_table(["Check", "Result", "Details"])
        rc.addWidget(self.chk_tbl); panels.addLayout(rc, 1)

        lay.addLayout(panels)
        lay.addWidget(lbl("Course Summary", 13, bold=True))
        self.course_tbl = make_table(
            ["Course", "Capacity", "Assigned TAs", "AvgUtil", "≥ k"])
        lay.addWidget(self.course_tbl)

        exp_row = QHBoxLayout()
        self.btn_exp = make_btn("  💾  Export to Excel  ", ACCENT2)
        self.btn_exp.setEnabled(False)
        self.btn_exp.clicked.connect(self._export)
        exp_row.addStretch(); exp_row.addWidget(self.btn_exp); exp_row.addStretch()
        lay.addLayout(exp_row)
        return w

    # ══════════════════════════════════════════════════════
    #  TAB 3 — GRAPH
    # ══════════════════════════════════════════════════════
    def _tab_graph(self):
        w = QWidget(); lay = QVBoxLayout(w); lay.setSpacing(10)

        # ── Controls ──
        ctrl = QHBoxLayout(); ctrl.setSpacing(16)

        self.chk_weights = QCheckBox("Show edge weights")
        self.chk_weights.setChecked(True)
        self.chk_weights.stateChanged.connect(self._refresh_graph)
        ctrl.addWidget(self.chk_weights)

        ctrl.addWidget(lbl("Weight:", 12, color=SUBTEXT))
        self.cmb_weight = QComboBox()
        self.cmb_weight.addItems([
            "TA Utility  (u)",
            "Course Utility  (v)",
            "Grade  (g)",
            "Both u & v"])
        self.cmb_weight.currentIndexChanged.connect(self._refresh_graph)
        ctrl.addWidget(self.cmb_weight)

        ctrl.addStretch()

        ctrl.addWidget(lbl("View:", 12, color=SUBTEXT))
        self.cmb_view = QComboBox()
        self.cmb_view.addItems([
            "After Matching  (show result)",
            "Before Matching  (graph only)"])
        self.cmb_view.currentIndexChanged.connect(self._refresh_graph)
        ctrl.addWidget(self.cmb_view)

        btn_ref = make_btn("  🔄 Refresh  ", ACCENT)
        btn_ref.setFixedHeight(36)
        btn_ref.clicked.connect(self._refresh_graph)
        ctrl.addWidget(btn_ref)

        lay.addLayout(ctrl)

        # ── Inline legend strip ──
        leg = QHBoxLayout(); leg.setSpacing(14)
        for color, text in [
            (COL_COURSE,     "Course node"),
            (COL_TA,         "TA node"),
            (COL_FORCED,     "Forced (Case 1)"),
            (COL_CYCLE,      "Cycle node"),
            (COL_UNASSIGNED, "Unassigned TA"),
            (COL_MATCHED,    "Matched edge ══"),
            (COL_EDGE,       "Unmatched  - -"),
        ]:
            dot = QLabel("●")
            dot.setStyleSheet(
                f"color:{color};font-size:16px;background:transparent;")
            leg.addWidget(dot)
            leg.addWidget(lbl(text, 11, color=SUBTEXT))
        leg.addStretch()
        lay.addLayout(leg)

        # ── Canvas ──
        self.canvas = GraphCanvas()
        lay.addWidget(self.canvas, stretch=1)

        # ── Info bar ──
        self.graph_info = QLabel(
            "Run the algorithm to populate the graph.")
        self.graph_info.setAlignment(Qt.AlignCenter)
        self.graph_info.setStyleSheet(
            f"color:{SUBTEXT};font-size:11px;"
            f"background:{CARD};border-radius:8px;"
            f"padding:8px;border:1px solid {BORDER};")
        lay.addWidget(self.graph_info)
        return w

    # ══════════════════════════════════════════════════════
    #  TAB 4 — LOG
    # ══════════════════════════════════════════════════════
    def _tab_log(self):
        w = QWidget(); lay = QVBoxLayout(w)
        lay.addWidget(lbl("Algorithm Execution Log", 14, bold=True))
        self.log_box = QTextEdit(); self.log_box.setReadOnly(True)
        self.log_box.setPlaceholderText(
            "Run the algorithm to see step-by-step log…")
        lay.addWidget(self.log_box)
        return w

    # ══════════════════════════════════════════════════════
    #  ACTIONS
    # ══════════════════════════════════════════════════════
    def _browse(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel Input", "",
            "Excel Files (*.xlsx *.xls);;All Files (*)")
        if path: self._load(path)

    def _load(self, path):
        try:
            courses, tas, caps, grade, ta_u, cu, k = read_excel(path)
            self._data = dict(
                courses=courses, tas=tas, capacities=caps,
                grade=grade, ta_utility=ta_u,
                course_utility=cu, k=k, path=path,
                matching=None, report=None,
                nc=None, nt=None, meta={})

            self.file_lbl.setText(f"✅  {os.path.basename(path)}")
            self.file_lbl.setStyleSheet(
                f"color:{GREEN};font-size:13px;background:transparent;")

            self._upd_sc("Courses",     f"{len(courses)}")
            self._upd_sc("TAs",         f"{len(tas)}")
            self._upd_sc("Grade",       f"{len(tas)}×{len(courses)}")
            self._upd_sc("TA Utility",  f"{len(tas)}×{len(courses)}")
            self._upd_sc("Course Util", f"{len(courses)}×{len(tas)}")
            self._upd_sc("k",           f"k={k}")

            self._show_previews(courses, tas, caps, grade, ta_u, cu)
            self.btn_run.setEnabled(True)
        except Exception as e:
            QMessageBox.critical(self, "Error reading file", str(e))

    def _show_previews(self, courses, tas, caps, grade, ta_u, cu):
        self.prev_tabs.clear(); self.prev_tabs.setVisible(True)

        t = make_table(["Course", "Capacity"])
        for c in courses: add_row(t, [c, caps[c]])
        self.prev_tabs.addTab(t, "Courses")

        t2 = make_table(["TA Name"])
        for ta in tas: add_row(t2, [ta])
        self.prev_tabs.addTab(t2, "TAs")

        for title, mat, rows, cols, accent in [
            ("Grade",       grade, tas,     courses, "#ffd700"),
            ("TA Utility",  ta_u,  tas,     courses, ACCENT2),
            ("Course Util", cu,    courses, tas,     ACCENT),
        ]:
            tw = QTableWidget(len(rows), len(cols))
            tw.setHorizontalHeaderLabels(cols)
            tw.setVerticalHeaderLabels(rows)
            tw.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            tw.setEditTriggers(QTableWidget.NoEditTriggers)
            for i, r in enumerate(rows):
                for j, c in enumerate(cols):
                    val  = mat.get((r, c), 0)
                    item = QTableWidgetItem(str(val))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setForeground(
                        QColor(accent if val != 0 else SUBTEXT))
                    tw.setItem(i, j, item)
            self.prev_tabs.addTab(tw, title)

    def _run(self):
        if not self._data: return
        self.btn_run.setEnabled(False)
        self.btn_run.setText("  ⏳  Running…")
        d = self._data

        matching, report, log, nc, nt, meta = solve(
            d['courses'], d['tas'], d['capacities'],
            d['grade'], d['ta_utility'], d['course_utility'], d['k'])

        d['matching'] = matching
        d['report']   = report
        d['nc']  = nc;  d['nt']   = nt
        d['meta'] = meta

        # log tab
        self.log_box.setPlainText("\n".join(log))

        # verdict banner
        if matching:
            self.verdict.setText("✅  VALID MEFE MATCHING FOUND!")
            self.verdict.setStyleSheet(
                f"background:#1a3a2a;color:{GREEN};"
                f"border-radius:10px;padding:14px;"
                f"font-size:16px;font-weight:700;"
                f"border:2px solid {GREEN};")
        else:
            self.verdict.setText("❌  NO VALID MEFE MATCHING EXISTS")
            self.verdict.setStyleSheet(
                f"background:#3a1a1a;color:{DANGER};"
                f"border-radius:10px;padding:14px;"
                f"font-size:16px;font-weight:700;"
                f"border:2px solid {DANGER};")

        # matching table
        self.match_tbl.setRowCount(0)
        if matching:
            for ta, c in sorted(matching.items()):
                add_row(self.match_tbl, [ta, c], [ACCENT2, ACCENT])
        for ta in d['tas']:
            if not matching or ta not in matching:
                add_row(self.match_tbl,
                        [ta, "— unassigned —"],
                        [SUBTEXT, SUBTEXT])

        # checks table
        self.chk_tbl.setRowCount(0)
        if report:
            for name, info in report.items():
                ok = info['passed']
                add_row(self.chk_tbl,
                        [name,
                         "✅ PASS" if ok else "❌ FAIL",
                         info['message']],
                        [None,
                         GREEN if ok else DANGER,
                         SUBTEXT])

        # course summary
        self.course_tbl.setRowCount(0)
        if matching and report:
            assigned  = defaultdict(list)
            for ta, c in matching.items(): assigned[c].append(ta)
            avg_utils = report.get('Satisfaction', {}).get('avg_utils', {})
            k = d['k']
            for c in d['courses']:
                tl  = sorted(assigned.get(c, []))
                avg = avg_utils.get(c, 0)
                ok  = avg >= k
                add_row(self.course_tbl,
                        [c, d['capacities'][c],
                         ", ".join(tl) or "—",
                         f"{avg:.2f}",
                         "YES" if ok else "NO"],
                        [ACCENT, None, ACCENT2,
                         GREEN if ok else DANGER,
                         GREEN if ok else DANGER])

        self.btn_run.setEnabled(True)
        self.btn_run.setText("  ▶  Run Algorithm  ")
        self.btn_exp.setEnabled(bool(matching))

        # draw graph and jump to tab
        self._refresh_graph()
        self.tabs.setCurrentIndex(2)

    def _refresh_graph(self):
        d = self._data
        if not d or not d.get('nc'): return

        wt_map  = ["TA Utility", "Course Utility", "Grade", "Both u & v"]
        wt_mode = wt_map[self.cmb_weight.currentIndex()]
        after   = self.cmb_view.currentIndex() == 0
        match   = d.get('matching') if after else {}

        self.canvas.draw_graph(
            d['courses'], d['tas'], d['capacities'],
            d['nc'], match or {},
            d.get('meta', {}),
            self.chk_weights.isChecked(),
            d['ta_utility'], d['course_utility'],
            d['grade'], wt_mode)

        parts = []
        for idx, m in d.get('meta', {}).items():
            cn = {'case1': 'Forced', 'case2_1': 'Tree',
                  'case2_2': 'Cycle',  'case2_3': 'Multi-cycle'
                  }.get(m['case'], '?')
            parts.append(
                f"Component {idx+1}: {cn}  "
                f"({len(m['courses'])} courses, {len(m['tas'])} TAs)")
        self.graph_info.setText(
            "   |   ".join(parts) if parts else
            "No graph data yet.")

    def _export(self):
        if not self._data.get('matching'): return
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Result", "mefe_result.xlsx",
            "Excel Files (*.xlsx)")
        if not path: return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb  = openpyxl.Workbook()
            ws  = wb.active; ws.title = "Matching"
            hf  = Font(bold=True, color="FFFFFF")
            hb  = PatternFill("solid", fgColor="2F5496")
            gf  = PatternFill("solid", fgColor="C6EFCE")
            rf  = PatternFill("solid", fgColor="FFC7CE")
            ctr = Alignment(horizontal="center")
            for col, h in enumerate(["TA", "Assigned Course"], 1):
                c = ws.cell(1, col, h); c.font = hf
                c.fill = hb; c.alignment = ctr
            for r, (ta, c) in enumerate(
                    sorted(self._data['matching'].items()), 2):
                ws.cell(r, 1, ta); ws.cell(r, 2, c)
            ws2 = wb.create_sheet("Validity")
            for col, h in enumerate(["Check","Result","Message"], 1):
                c = ws2.cell(1, col, h); c.font = hf; c.fill = hb
            for r, (name, info) in enumerate(
                    (self._data.get('report') or {}).items(), 2):
                ws2.cell(r, 1, name)
                res = ws2.cell(r, 2,
                               "PASS" if info['passed'] else "FAIL")
                res.fill = gf if info['passed'] else rf
                ws2.cell(r, 3, info['message'])
            wb.save(path)
            QMessageBox.information(self, "Saved", f"Saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))

    def _make_sample(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Sample", "mefe_sample.xlsx",
            "Excel Files (*.xlsx)")
        if not path: return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            wb  = openpyxl.Workbook()
            hf  = Font(bold=True, color='FFFFFF')
            hb  = PatternFill('solid', fgColor='2F5496')
            ctr = Alignment(horizontal='center')
            def hdr(c): c.font = hf; c.fill = hb; c.alignment = ctr
            ws = wb.active; ws.title = 'Courses'
            hdr(ws.cell(1,1,'Course Name')); hdr(ws.cell(1,2,'Capacity'))
            for i,(n,c) in enumerate([('x1',2),('x2',1)],2):
                ws.cell(i,1,n); ws.cell(i,2,c)
            ws2 = wb.create_sheet('TAs'); hdr(ws2.cell(1,1,'TA Name'))
            for i,t in enumerate(['t1','t2','t3','t4'],2): ws2.cell(i,1,t)
            ws3 = wb.create_sheet('Grade_g'); ws3.cell(1,1,'')
            for j,c in enumerate(['x1','x2'],2): hdr(ws3.cell(1,j,c))
            for i,(ta,g1,g2) in enumerate(
                    [('t1',3,0),('t2',8,0),('t3',5,4),('t4',0,9)],2):
                ws3.cell(i,1,ta); ws3.cell(i,2,g1); ws3.cell(i,3,g2)
            ws4 = wb.create_sheet('TA_Utility_u'); ws4.cell(1,1,'')
            for j,c in enumerate(['x1','x2'],2): hdr(ws4.cell(1,j,c))
            for i,(ta,u1,u2) in enumerate(
                    [('t1',0,0),('t2',3,0),('t3',1,3),('t4',0,2)],2):
                ws4.cell(i,1,ta); ws4.cell(i,2,u1); ws4.cell(i,3,u2)
            ws5 = wb.create_sheet('Course_Utility_v'); ws5.cell(1,1,'')
            for j,ta in enumerate(['t1','t2','t3','t4'],2):
                hdr(ws5.cell(1,j,ta))
            for i,(c,vals) in enumerate(
                    [('x1',[0,3,2,0]),('x2',[0,0,5,4])],2):
                ws5.cell(i,1,c)
                for j,v in enumerate(vals,2): ws5.cell(i,j,v)
            ws6 = wb.create_sheet('Threshold_k')
            hdr(ws6.cell(1,1,'k')); ws6.cell(2,1,2)
            wb.save(path)
            QMessageBox.information(self,"Sample Created",
                f"Saved to:\n{path}\n\nNow upload it using Browse File.")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))


# ═══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    win = MEFEApp()
    win.show()
    sys.exit(app.exec_())