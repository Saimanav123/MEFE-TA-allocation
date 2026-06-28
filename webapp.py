import os
import uuid
import itertools
from io import BytesIO
from pathlib import Path
from collections import defaultdict, deque

import pandas as pd
import networkx as nx
import openpyxl
from flask import Flask, jsonify, render_template, request, send_file


BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)


def _load_backend_namespace():
    """Load the matching backend from main.py without importing the PyQt GUI."""
    source = (BASE_DIR / "main.py").read_text(encoding="utf-8", errors="ignore")
    start = source.index("THEOREM_LABELS = [")
    end = source.index("class SmartSelectionPanel(QWidget):")
    snippet = source[start:end]

    namespace = {
        "defaultdict": defaultdict,
        "deque": deque,
        "itertools": itertools,
        "pd": pd,
        "nx": nx,
        "GREEN": "#2ed573",
        "ACCENT": "#6c63ff",
        "ACCENT2": "#00d4aa",
        "DANGER": "#ff4757",
    }
    exec(snippet, namespace)
    return namespace


BACKEND = _load_backend_namespace()
read_excel = BACKEND["read_excel"]
analyze_matching_data = BACKEND["analyze_matching_data"]
execute_strategy = BACKEND["execute_strategy"]
CUSTOM_BACKEND_CHOICES = BACKEND["CUSTOM_BACKEND_CHOICES"]

STRATEGY_CARDS = [
    {
        "id": "auto",
        "title": "Auto Select Best Method",
        "icon": "AI",
        "badge": "Recommended",
        "tone": "success",
        "description": "The system studies your dataset and quietly picks the best path for you.",
        "tooltip": "Best for most users and the safest first choice.",
    },
    {
        "id": "fast",
        "title": "Fast Matching",
        "icon": "FAST",
        "badge": "Speed",
        "tone": "info",
        "description": "Built for larger files or simple TA choices where quick results matter most.",
        "tooltip": "Use this when you want the shortest path to a good result.",
    },
    {
        "id": "accurate",
        "title": "High Accuracy Matching",
        "icon": "HQ",
        "badge": "Quality",
        "tone": "primary",
        "description": "Focuses on stronger preference alignment when your input structure supports it.",
        "tooltip": "Best when the recommendation says the data supports quality-first matching.",
    },
    {
        "id": "compare",
        "title": "Try All Methods",
        "icon": "ALL",
        "badge": "Heavy",
        "tone": "warning",
        "description": "Runs multiple internal methods, compares outcomes, and keeps the strongest result.",
        "tooltip": "Useful for smaller datasets when you want confidence through comparison.",
    },
    {
        "id": "custom",
        "title": "Custom Selection",
        "icon": "ADV",
        "badge": "Advanced",
        "tone": "muted",
        "description": "Lets you override the recommendation and choose the internal engine manually.",
        "tooltip": "Good for advanced experimentation when you know you want a specific backend path.",
    },
]

SESSION_STORE = {}

app = Flask(__name__)


def _serialize_analysis(analysis):
    return {
        "num_courses": analysis["num_courses"],
        "num_tas": analysis["num_tas"],
        "total_slots": analysis["total_slots"],
        "positive_pairs": analysis["positive_pairs"],
        "avg_ta_degree": round(analysis["avg_ta_degree"], 2),
        "avg_course_degree": round(analysis["avg_course_degree"], 2),
        "density": round(analysis["density"] * 100, 1),
        "recommended_strategy": analysis["recommended_strategy"],
        "recommended_backend": analysis["recommended_backend"],
        "recommendation_text": analysis["recommendation_text"],
        "why_lines": list(analysis["why_lines"]),
        "insights": list(analysis["insights"]),
        "large_data": analysis["large_data"],
        "small_data": analysis["small_data"],
        "ta_degree_le_one": analysis["ta_degree_le_one"],
        "limited_choices": analysis["limited_choices"],
        "all_capacity_one": analysis["all_capacity_one"],
    }


def _build_course_summary(data, matching, report):
    assigned = defaultdict(list)
    for ta, course in (matching or {}).items():
        assigned[course].append(ta)

    avg_utils = (report or {}).get("Satisfaction", {}).get("avg_utils", {})
    rows = []
    for course in data["courses"]:
        avg = float(avg_utils.get(course, 0))
        rows.append(
            {
                "course": course,
                "capacity": data["capacities"][course],
                "assigned_tas": ", ".join(sorted(assigned.get(course, []))) or "Unassigned",
                "avg_util": round(avg, 2),
                "meets_threshold": avg >= data["k"],
            }
        )
    return rows


def _extract_failure_reason(result, summary_text):
    lines = []
    if result:
        lines.extend(result.get("log") or [])
    if summary_text:
        lines.extend(summary_text.splitlines())

    priority_markers = [
        "does NOT apply",
        "requires",
        "violating",
        "No valid",
        "No feasible",
        "not enough",
        "failed",
        "violated",
        "No stable",
        "No result",
    ]
    for line in lines:
        cleaned = " ".join(str(line).split())
        if not cleaned:
            continue
        for marker in priority_markers:
            if marker.lower() in cleaned.lower():
                return cleaned

    for line in lines:
        cleaned = " ".join(str(line).split())
        if cleaned:
            return cleaned
    return "This selection could not produce a valid matching for the uploaded data."


def _fallback_failed_checks(reason):
    return [
        {"name": "Feasibility", "passed": False, "message": reason},
        {"name": "Satisfaction", "passed": False, "message": reason},
        {"name": "Envy-Freeness", "passed": False, "message": reason},
    ]


def _serialize_result(data, strategy_id, result, summary_text):
    if not result:
        return {
            "ok": False,
            "strategy": strategy_id,
            "engine_label": "No engine selected",
            "verdict": "No valid matching could be completed with the current data.",
            "matching_rows": [],
            "checks": _fallback_failed_checks(
                _extract_failure_reason(result, summary_text)
            ),
            "course_summary": [],
            "summary_text": summary_text,
        }

    matching = result["matching"]
    report = result["report"]
    checks = []
    for name, info in (report or {}).items():
        checks.append(
            {
                "name": name,
                "passed": bool(info["passed"]),
                "message": info["message"],
            }
        )

    if not checks:
        checks = _fallback_failed_checks(_extract_failure_reason(result, summary_text))

    matching_rows = []
    for ta in data["tas"]:
        matching_rows.append(
            {
                "ta": ta,
                "course": matching.get(ta, "Unassigned"),
                "assigned": ta in matching,
            }
        )

    all_checks_passed = bool(checks) and all(item["passed"] for item in checks)
    verdict = (
        f"Matching complete using {result['backend_label']}"
        if matching and all_checks_passed
        else _extract_failure_reason(result, summary_text)
    )

    return {
        "ok": bool(matching) and all_checks_passed,
        "strategy": strategy_id,
        "engine_label": result["backend_label"],
        "verdict": verdict,
        "matching_rows": matching_rows,
        "checks": checks,
        "course_summary": _build_course_summary(data, matching, report) if matching else [],
        "summary_text": summary_text,
    }


def _build_export_workbook(data, latest_result):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matching"
    ws.append(["TA", "Assigned Course"])
    for row in latest_result["matching_rows"]:
        ws.append([row["ta"], row["course"]])

    ws2 = wb.create_sheet("Validity")
    ws2.append(["Check", "Result", "Message"])
    for check in latest_result["checks"]:
        ws2.append([check["name"], "PASS" if check["passed"] else "FAIL", check["message"]])

    ws3 = wb.create_sheet("Course Summary")
    ws3.append(["Course", "Capacity", "Assigned TAs", "AvgUtil", ">= k"])
    for row in latest_result["course_summary"]:
        ws3.append([
            row["course"],
            row["capacity"],
            row["assigned_tas"],
            row["avg_util"],
            "YES" if row["meets_threshold"] else "NO",
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@app.get("/")
def index():
    return render_template("index.html")


@app.post("/api/analyze")
def analyze():
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "Please choose an Excel file first."}), 400

    session_id = uuid.uuid4().hex
    ext = Path(file.filename).suffix or ".xlsx"
    saved_path = UPLOAD_DIR / f"{session_id}{ext}"
    file.save(saved_path)

    try:
        courses, tas, capacities, grade, ta_utility, course_utility, k = read_excel(saved_path)
        data = {
            "path": str(saved_path),
            "filename": file.filename,
            "courses": courses,
            "tas": tas,
            "capacities": capacities,
            "grade": grade,
            "ta_utility": ta_utility,
            "course_utility": course_utility,
            "k": k,
        }
        analysis = analyze_matching_data(courses, tas, capacities, grade, ta_utility, course_utility)
        SESSION_STORE[session_id] = {
            "data": data,
            "analysis": analysis,
            "result": None,
        }
    except Exception as exc:
        if saved_path.exists():
            saved_path.unlink(missing_ok=True)
        return jsonify({"error": f"Could not read the file: {exc}"}), 400

    return jsonify(
        {
            "session_id": session_id,
            "filename": file.filename,
            "analysis": _serialize_analysis(analysis),
            "strategies": STRATEGY_CARDS,
            "custom_choices": [
                {"label": label, "value": value} for label, value in CUSTOM_BACKEND_CHOICES
            ],
        }
    )


@app.post("/api/run")
def run_matching():
    payload = request.get_json(silent=True) or {}
    session_id = payload.get("session_id")
    strategy = payload.get("strategy", "auto")
    custom_backend = payload.get("custom_backend")

    if not session_id or session_id not in SESSION_STORE:
        return jsonify({"error": "Your session was not found. Please upload the file again."}), 400

    entry = SESSION_STORE[session_id]
    data = entry["data"]
    analysis = entry["analysis"]

    result, summary_text = execute_strategy(data, strategy, analysis, custom_backend)
    serialized = _serialize_result(data, strategy, result, summary_text)
    entry["result"] = serialized
    return jsonify(serialized)


@app.get("/api/export/<session_id>")
def export_result(session_id):
    entry = SESSION_STORE.get(session_id)
    if not entry or not entry.get("result"):
        return jsonify({"error": "No result is ready to export yet."}), 404

    workbook = _build_export_workbook(entry["data"], entry["result"])
    filename = f"mefe_result_{session_id[:8]}.xlsx"
    return send_file(
        workbook,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(debug=True, port=5000)
